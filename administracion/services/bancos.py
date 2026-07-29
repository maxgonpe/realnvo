"""
B001–B021 — Banco y conciliación.

Activos: B001 bancos (+ clave cartola), B002 cuentas, B004 plantillas.
Pendiente:
- Importación cartola (B003) + duplicados (B005)
- Aplicaciones múltiples (B009–B011)
- Reglas / sugerencias de cruce (B012–B013)
- Reversas y cierre de período (B016–B017)
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Prefetch
from django.utils import timezone

from ..crypto import cifrar_texto, descifrar_texto
from ..models import (
    Banco,
    CampoMapeoCartola,
    CuentaBancaria,
    ImportacionCartola,
    PlantillaMapeoCartola,
)

CAMPOS_OBLIGATORIOS_BASE = {
    CampoMapeoCartola.CampoDestino.FECHA_OPERACION,
    CampoMapeoCartola.CampoDestino.DESCRIPCION,
}


# ---------------------------------------------------------------------------
# B001 — Bancos + clave de cartola
# ---------------------------------------------------------------------------


def normalizar_nombre_banco(nombre: str) -> str:
    texto = re.sub(r"\s+", " ", (nombre or "").strip())
    return texto


def normalizar_codigo_banco(codigo: str) -> str:
    return (codigo or "").strip().upper()


def _validar_nombre_unico(nombre: str, exclude_pk: int | None = None) -> None:
    qs = Banco.objects.all()
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    for existente in qs.only("id", "nombre"):
        if existente.nombre.casefold() == nombre.casefold():
            raise ValidationError(
                {"nombre": f"Ya existe el banco «{existente.nombre}»."}
            )


def _validar_codigo_unico(codigo: str, exclude_pk: int | None = None) -> None:
    if not codigo:
        return
    qs = Banco.objects.filter(codigo=codigo)
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    if qs.exists():
        raise ValidationError({"codigo": "Ya existe un banco con este código."})


@transaction.atomic
def crear_banco(*, datos: dict, usuario) -> Banco:
    nombre = normalizar_nombre_banco(datos.get("nombre") or "")
    codigo = normalizar_codigo_banco(datos.get("codigo") or "")
    if not nombre:
        raise ValidationError({"nombre": "El nombre es obligatorio."})
    _validar_nombre_unico(nombre)
    _validar_codigo_unico(codigo)

    banco = Banco(
        nombre=nombre,
        codigo=codigo,
        activo=bool(datos.get("activo", True)),
        creado_por=usuario,
        actualizado_por=usuario,
    )
    banco.full_clean()
    banco.save()

    clave = (datos.get("clave_cartola") or "").strip()
    if clave:
        definir_clave_cartola(banco=banco, clave=clave, usuario=usuario)
        banco.refresh_from_db()
    return banco


@transaction.atomic
def actualizar_banco(*, banco: Banco, datos: dict, usuario) -> Banco:
    nombre = normalizar_nombre_banco(datos.get("nombre") or "")
    codigo = normalizar_codigo_banco(datos.get("codigo") or "")
    activo = bool(datos.get("activo", banco.activo))
    if not nombre:
        raise ValidationError({"nombre": "El nombre es obligatorio."})
    _validar_nombre_unico(nombre, exclude_pk=banco.pk)
    _validar_codigo_unico(codigo, exclude_pk=banco.pk)

    if not activo and banco.cuentas.filter(activa=True).exists():
        raise ValidationError(
            "No se puede desactivar un banco con cuentas activas. "
            "Desactive primero las cuentas (B002)."
        )

    banco.nombre = nombre
    banco.codigo = codigo
    banco.activo = activo
    banco.actualizado_por = usuario
    banco.full_clean()
    banco.save()
    return banco


@transaction.atomic
def activar_banco(*, banco: Banco, usuario) -> Banco:
    banco.activo = True
    banco.actualizado_por = usuario
    banco.save(update_fields=["activo", "actualizado_por", "actualizado_en"])
    return banco


@transaction.atomic
def desactivar_banco(*, banco: Banco, usuario) -> Banco:
    if banco.cuentas.filter(activa=True).exists():
        raise ValidationError(
            "No se puede desactivar un banco con cuentas activas."
        )
    banco.activo = False
    banco.actualizado_por = usuario
    banco.save(update_fields=["activo", "actualizado_por", "actualizado_en"])
    return banco


@transaction.atomic
def definir_clave_cartola(*, banco: Banco, clave: str, usuario) -> Banco:
    """Guarda la clave PDF cifrada. Vacío = no cambia."""
    clave = (clave or "").strip()
    if not clave:
        raise ValidationError({"clave_cartola": "Indique la nueva clave."})
    banco.clave_cartola_cifrada = cifrar_texto(clave)
    banco.clave_cartola_actualizada_en = timezone.now()
    banco.actualizado_por = usuario
    banco.save(
        update_fields=[
            "clave_cartola_cifrada",
            "clave_cartola_actualizada_en",
            "actualizado_por",
            "actualizado_en",
        ]
    )
    return banco


@transaction.atomic
def limpiar_clave_cartola(*, banco: Banco, usuario) -> Banco:
    banco.clave_cartola_cifrada = ""
    banco.clave_cartola_actualizada_en = timezone.now()
    banco.actualizado_por = usuario
    banco.save(
        update_fields=[
            "clave_cartola_cifrada",
            "clave_cartola_actualizada_en",
            "actualizado_por",
            "actualizado_en",
        ]
    )
    return banco


def obtener_clave_cartola(banco: Banco) -> str:
    """Solo para servicios de importación (B003). No exponer en templates."""
    if not banco.clave_cartola_cifrada:
        raise ValidationError(
            f"El banco «{banco.nombre}» no tiene clave de cartola configurada."
        )
    try:
        return descifrar_texto(banco.clave_cartola_cifrada)
    except ValueError as exc:
        raise ValidationError(str(exc)) from exc


def verificar_clave_contra_pdf(*, banco: Banco, archivo) -> bool:
    """Prueba que la clave guardada abre el PDF. No deja la clave en logs."""
    from pypdf import PdfReader

    clave = obtener_clave_cartola(banco)
    reader = PdfReader(archivo)
    if not reader.is_encrypted:
        return True
    return bool(reader.decrypt(clave))


# ---------------------------------------------------------------------------
# B002 — Cuentas bancarias
# ---------------------------------------------------------------------------


def normalizar_numero_cuenta(numero: str) -> str:
    """Quita espacios y separadores visuales; conserva dígitos, letras y X de enmascarado."""
    bruto = (numero or "").strip().upper()
    return re.sub(r"[\s.\-_/]", "", bruto)


def _validar_numero_unico(
    banco: Banco, numero: str, exclude_pk: int | None = None
) -> None:
    qs = CuentaBancaria.objects.filter(banco=banco, numero_cuenta=numero)
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    if qs.exists():
        raise ValidationError(
            {"numero_cuenta": "Ya existe esa cuenta en este banco."}
        )


@transaction.atomic
def crear_cuenta(*, datos: dict, usuario) -> CuentaBancaria:
    banco = datos.get("banco")
    if not banco:
        raise ValidationError({"banco": "Seleccione un banco."})
    if not banco.activo:
        raise ValidationError({"banco": "El banco debe estar activo."})

    numero = normalizar_numero_cuenta(datos.get("numero_cuenta") or "")
    if not numero:
        raise ValidationError({"numero_cuenta": "El número de cuenta es obligatorio."})
    _validar_numero_unico(banco, numero)

    alias = re.sub(r"\s+", " ", (datos.get("nombre") or "").strip())
    if not alias:
        raise ValidationError({"nombre": "El alias de la cuenta es obligatorio."})

    cuenta = CuentaBancaria(
        banco=banco,
        nombre=alias,
        numero_cuenta=numero,
        tipo_cuenta=datos.get("tipo_cuenta") or CuentaBancaria.TipoCuenta.CORRIENTE,
        moneda=(datos.get("moneda") or "CLP").strip().upper()[:3] or "CLP",
        titular=re.sub(r"\s+", " ", (datos.get("titular") or "").strip()),
        rut_titular=(datos.get("rut_titular") or "").strip(),
        activa=bool(datos.get("activa", True)),
        observaciones=(datos.get("observaciones") or "").strip(),
        creado_por=usuario,
        actualizado_por=usuario,
    )
    cuenta.full_clean()
    cuenta.save()
    return cuenta


@transaction.atomic
def actualizar_cuenta(
    *, cuenta: CuentaBancaria, datos: dict, usuario
) -> CuentaBancaria:
    banco = datos.get("banco") or cuenta.banco
    if not banco.activo and banco.pk != cuenta.banco_id:
        raise ValidationError({"banco": "El banco destino debe estar activo."})

    numero = normalizar_numero_cuenta(datos.get("numero_cuenta") or "")
    if not numero:
        raise ValidationError({"numero_cuenta": "El número de cuenta es obligatorio."})
    _validar_numero_unico(banco, numero, exclude_pk=cuenta.pk)

    alias = re.sub(r"\s+", " ", (datos.get("nombre") or "").strip())
    if not alias:
        raise ValidationError({"nombre": "El alias de la cuenta es obligatorio."})

    activa = bool(datos.get("activa", cuenta.activa))
    if not activa:
        pendientes = cuenta.importaciones.filter(estado="PENDIENTE").exists()
        if pendientes:
            raise ValidationError(
                "No se puede desactivar: hay importaciones pendientes."
            )

    cuenta.banco = banco
    cuenta.nombre = alias
    cuenta.numero_cuenta = numero
    cuenta.tipo_cuenta = (
        datos.get("tipo_cuenta") or cuenta.tipo_cuenta
    )
    cuenta.moneda = (datos.get("moneda") or "CLP").strip().upper()[:3] or "CLP"
    cuenta.titular = re.sub(r"\s+", " ", (datos.get("titular") or "").strip())
    cuenta.rut_titular = (datos.get("rut_titular") or "").strip()
    cuenta.activa = activa
    cuenta.observaciones = (datos.get("observaciones") or "").strip()
    cuenta.actualizado_por = usuario
    cuenta.full_clean()
    cuenta.save()
    return cuenta


@transaction.atomic
def activar_cuenta(*, cuenta: CuentaBancaria, usuario) -> CuentaBancaria:
    if not cuenta.banco.activo:
        raise ValidationError("Reactive primero el banco (B001).")
    cuenta.activa = True
    cuenta.actualizado_por = usuario
    cuenta.save(update_fields=["activa", "actualizado_por", "actualizado_en"])
    return cuenta


@transaction.atomic
def desactivar_cuenta(*, cuenta: CuentaBancaria, usuario) -> CuentaBancaria:
    if cuenta.importaciones.filter(estado="PENDIENTE").exists():
        raise ValidationError("No se puede desactivar: hay importaciones pendientes.")
    cuenta.activa = False
    cuenta.actualizado_por = usuario
    cuenta.save(update_fields=["activa", "actualizado_por", "actualizado_en"])
    return cuenta


@transaction.atomic
def crear_cuenta_desde_cartola_pdf(
    *, banco: Banco, archivo, usuario, forzar: bool = False
) -> tuple[CuentaBancaria, dict]:
    """
    Lee metadatos del PDF y crea (o actualiza) la cuenta.
    Devuelve (cuenta, metadatos).
    """
    from .cartola_pdf import leer_metadatos_cartola_pdf

    meta = leer_metadatos_cartola_pdf(banco=banco, archivo=archivo)
    numero = normalizar_numero_cuenta(meta.get("numero_cuenta") or "")
    if not numero:
        raise ValidationError(
            "No se pudo detectar el número de cuenta en el PDF. "
            "Cree la cuenta manualmente."
        )

    existente = CuentaBancaria.objects.filter(
        banco=banco, numero_cuenta=numero
    ).first()
    datos = {
        "banco": banco,
        "nombre": meta.get("nombre") or f"Cuenta {banco.nombre}",
        "numero_cuenta": numero,
        "tipo_cuenta": meta.get("tipo_cuenta") or CuentaBancaria.TipoCuenta.CORRIENTE,
        "moneda": meta.get("moneda") or "CLP",
        "titular": meta.get("titular") or "",
        "rut_titular": "",
        "activa": True,
        "observaciones": (
            f"Importado desde cartola PDF. "
            f"Período {meta.get('periodo_desde')} → {meta.get('periodo_hasta')}. "
            f"Saldo inicial detectado: {meta.get('saldo_inicial') or '—'}."
        ),
    }
    if existente and not forzar:
        cuenta = actualizar_cuenta(cuenta=existente, datos=datos, usuario=usuario)
    elif existente and forzar:
        cuenta = actualizar_cuenta(cuenta=existente, datos=datos, usuario=usuario)
    else:
        cuenta = crear_cuenta(datos=datos, usuario=usuario)
    return cuenta, meta


# ---------------------------------------------------------------------------
# B004 — Plantillas de mapeo
# ---------------------------------------------------------------------------



def _destinos(campos: list[dict] | list[CampoMapeoCartola]) -> set[str]:
    resultado = set()
    for c in campos:
        if isinstance(c, dict):
            if c.get("DELETE"):
                continue
            destino = c.get("campo_destino")
        else:
            destino = c.campo_destino
        if destino:
            resultado.add(destino)
    return resultado


def validar_esquema_monetario(campos: list[dict] | list[CampoMapeoCartola]) -> str:
    """
    Devuelve el esquema detectado: A (tipo+monto), B (monto con signo) o C (cargo/abono).
    """
    destinos = _destinos(campos)
    Destino = CampoMapeoCartola.CampoDestino

    tiene_a = Destino.TIPO in destinos and Destino.MONTO in destinos
    tiene_b = Destino.MONTO in destinos and Destino.TIPO not in destinos
    tiene_c = Destino.MONTO_INGRESO in destinos and Destino.MONTO_EGRESO in destinos

    esquemas = []
    if tiene_a:
        esquemas.append("A")
    if tiene_b and not tiene_a:
        esquemas.append("B")
    if tiene_c:
        esquemas.append("C")

    if len(esquemas) != 1:
        raise ValidationError(
            "Debe definir exactamente un esquema monetario: "
            "(A) tipo + monto, (B) monto con signo, o (C) cargo + abono."
        )
    return esquemas[0]


def validar_plantilla(
    *,
    plantilla: PlantillaMapeoCartola | None = None,
    campos: list[dict] | list[CampoMapeoCartola] | None = None,
) -> dict:
    """Valida reglas B004. No persiste."""
    if campos is None:
        if plantilla is None:
            raise ValidationError("Se requiere plantilla o campos.")
        campos = list(plantilla.campos.all())

    destinos = _destinos(campos)
    faltantes = [
        CampoMapeoCartola.CampoDestino(c).label
        for c in CAMPOS_OBLIGATORIOS_BASE
        if c not in destinos
    ]
    if faltantes:
        raise ValidationError(
            f"Faltan campos obligatorios: {', '.join(faltantes)}."
        )

    if len(destinos) != len(
        [d for d in destinos]
    ):  # pragma: no cover - set ya deduplica
        pass

    vistos: set[str] = set()
    for c in campos:
        destino = c.get("campo_destino") if isinstance(c, dict) else c.campo_destino
        if isinstance(c, dict) and c.get("DELETE"):
            continue
        if not destino:
            continue
        if destino in vistos:
            raise ValidationError(
                f"El campo destino «{destino}» está repetido en la plantilla."
            )
        vistos.add(destino)
        columna = (
            (c.get("columna_origen") or "").strip()
            if isinstance(c, dict)
            else (c.columna_origen or "").strip()
        )
        if not columna:
            raise ValidationError(
                f"La columna origen de «{destino}» no puede estar vacía."
            )

    esquema = validar_esquema_monetario(campos)
    return {"esquema": esquema, "destinos": sorted(destinos)}


def _validar_cuenta_del_banco(datos: dict) -> None:
    cuenta = datos.get("cuenta_bancaria")
    banco = datos.get("banco")
    if cuenta and banco and cuenta.banco_id != banco.pk:
        raise ValidationError(
            {"cuenta_bancaria": "La cuenta debe pertenecer al banco seleccionado."}
        )


def _guardar_campos(
    *, plantilla: PlantillaMapeoCartola, campos: list[dict], usuario
) -> None:
    plantilla.campos.all().delete()
    for idx, item in enumerate(campos):
        if item.get("DELETE"):
            continue
        destino = item.get("campo_destino")
        columna = (item.get("columna_origen") or "").strip()
        if not destino or not columna:
            continue
        CampoMapeoCartola.objects.create(
            plantilla=plantilla,
            campo_destino=destino,
            columna_origen=columna,
            obligatorio=bool(item.get("obligatorio", False)),
            valor_defecto=(item.get("valor_defecto") or "").strip(),
            orden=item.get("orden", idx),
            creado_por=usuario,
            actualizado_por=usuario,
        )


def _aplicar_datos_plantilla(
    plantilla: PlantillaMapeoCartola, datos: dict, usuario
) -> None:
    plantilla.banco = datos["banco"]
    plantilla.cuenta_bancaria = datos.get("cuenta_bancaria")
    plantilla.nombre = (datos.get("nombre") or "").strip()
    plantilla.formato_archivo = datos["formato_archivo"]
    plantilla.parser_codigo = datos.get("parser_codigo") or PlantillaMapeoCartola.ParserCodigo.GENERICO
    plantilla.nombre_hoja = (datos.get("nombre_hoja") or "").strip()
    plantilla.fila_encabezado = int(datos.get("fila_encabezado") or 1)
    plantilla.fila_inicio_datos = int(datos.get("fila_inicio_datos") or 2)
    plantilla.separador_csv = datos.get("separador_csv") or ";"
    plantilla.codificacion = datos.get("codificacion") or "utf-8-sig"
    plantilla.formato_fecha = datos.get("formato_fecha") or "%d/%m/%Y"
    plantilla.fecha_sin_anio = bool(datos.get("fecha_sin_anio", False))
    # Vacío = montos enteros (sin decimales)
    plantilla.separador_decimal = datos.get("separador_decimal") or ""
    plantilla.separador_miles = datos.get("separador_miles") or "."
    plantilla.simbolo_moneda = (datos.get("simbolo_moneda") or "$").strip()
    plantilla.identificador_saldo_inicial = (
        datos.get("identificador_saldo_inicial") or "SALDO INICIAL"
    )
    plantilla.identificador_saldo_final = (
        datos.get("identificador_saldo_final") or "SALDO FINAL"
    )
    plantilla.ignorar_filas_vacias = bool(datos.get("ignorar_filas_vacias", True))
    plantilla.activa = bool(datos.get("activa", True))
    plantilla.observaciones = (datos.get("observaciones") or "").strip()
    plantilla.actualizado_por = usuario
    if not plantilla.nombre:
        raise ValidationError({"nombre": "El nombre es obligatorio."})


def plantilla_fue_usada_en_importacion(plantilla: PlantillaMapeoCartola) -> bool:
    return plantilla.importaciones.filter(
        estado=ImportacionCartola.Estado.PROCESADA
    ).exists()


@transaction.atomic
def crear_plantilla(*, datos: dict, campos: list[dict], usuario) -> PlantillaMapeoCartola:
    datos = dict(datos)
    _validar_cuenta_del_banco(datos)
    es_pdf = datos.get("formato_archivo") == PlantillaMapeoCartola.FormatoArchivo.PDF
    if not es_pdf:
        validar_plantilla(campos=campos)

    plantilla = PlantillaMapeoCartola(creado_por=usuario)
    _aplicar_datos_plantilla(plantilla, datos, usuario)
    plantilla.version = 1
    plantilla.full_clean()
    plantilla.save()
    if not es_pdf:
        _guardar_campos(plantilla=plantilla, campos=campos, usuario=usuario)
    return plantilla


@transaction.atomic
def actualizar_plantilla(
    *,
    plantilla: PlantillaMapeoCartola,
    datos: dict,
    campos: list[dict],
    usuario,
) -> PlantillaMapeoCartola:
    """
    Si la plantilla ya fue usada en importaciones PROCESADAS, crea una nueva versión
    y desactiva la anterior (no destruye el histórico).
    """
    datos = dict(datos)
    _validar_cuenta_del_banco(datos)
    es_pdf = datos.get("formato_archivo") == PlantillaMapeoCartola.FormatoArchivo.PDF
    if not es_pdf:
        validar_plantilla(campos=campos)

    if plantilla_fue_usada_en_importacion(plantilla):
        plantilla.activa = False
        plantilla.actualizado_por = usuario
        plantilla.save(update_fields=["activa", "actualizado_por", "actualizado_en"])

        nueva = PlantillaMapeoCartola(
            creado_por=usuario,
            version=plantilla.version + 1,
        )
        _aplicar_datos_plantilla(nueva, datos, usuario)
        nueva.activa = True
        nueva.full_clean()
        nueva.save()
        if not es_pdf:
            _guardar_campos(plantilla=nueva, campos=campos, usuario=usuario)
        return nueva

    _aplicar_datos_plantilla(plantilla, datos, usuario)
    plantilla.full_clean()
    plantilla.save()
    if not es_pdf:
        _guardar_campos(plantilla=plantilla, campos=campos, usuario=usuario)
    return plantilla


@transaction.atomic
def activar_plantilla(
    *, plantilla: PlantillaMapeoCartola, usuario
) -> PlantillaMapeoCartola:
    validar_plantilla(plantilla=plantilla)
    plantilla.activa = True
    plantilla.actualizado_por = usuario
    plantilla.save(update_fields=["activa", "actualizado_por", "actualizado_en"])
    return plantilla


@transaction.atomic
def desactivar_plantilla(
    *, plantilla: PlantillaMapeoCartola, usuario
) -> PlantillaMapeoCartola:
    plantilla.activa = False
    plantilla.actualizado_por = usuario
    plantilla.save(update_fields=["activa", "actualizado_por", "actualizado_en"])
    return plantilla


def _parse_monto(raw: str, plantilla: PlantillaMapeoCartola) -> Decimal | None:
    if raw is None:
        return None
    texto = str(raw).strip()
    if not texto:
        return None
    texto = texto.replace(plantilla.separador_miles, "").replace(
        plantilla.separador_decimal, "."
    )
    texto = texto.replace("$", "").replace(" ", "")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValidationError(f"Monto inválido: {raw!r}") from exc


def _leer_filas_csv(archivo, plantilla: PlantillaMapeoCartola) -> tuple[list[str], list[dict]]:
    raw = archivo.read()
    if isinstance(raw, bytes):
        texto = raw.decode(plantilla.codificacion or "utf-8-sig", errors="replace")
    else:
        texto = raw
    reader = csv.reader(io.StringIO(texto), delimiter=plantilla.separador_csv or ";")
    filas = list(reader)
    return _filas_a_dicts(filas, plantilla)


def _leer_filas_xlsx(archivo, plantilla: PlantillaMapeoCartola) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(archivo, read_only=True, data_only=True)
    if plantilla.nombre_hoja:
        if plantilla.nombre_hoja not in wb.sheetnames:
            raise ValidationError(
                f"La hoja «{plantilla.nombre_hoja}» no existe. "
                f"Hojas: {', '.join(wb.sheetnames)}"
            )
        ws = wb[plantilla.nombre_hoja]
    else:
        ws = wb.active
    filas = [
        ["" if c is None else str(c) for c in row]
        for row in ws.iter_rows(values_only=True)
    ]
    return _filas_a_dicts(filas, plantilla)


def _filas_a_dicts(
    filas: list[list[Any]], plantilla: PlantillaMapeoCartola
) -> tuple[list[str], list[dict]]:
    idx_enc = max(plantilla.fila_encabezado - 1, 0)
    idx_datos = max(plantilla.fila_inicio_datos - 1, 0)
    if idx_enc >= len(filas):
        raise ValidationError("La fila de encabezado no existe en el archivo.")
    encabezados = [str(c).strip() for c in filas[idx_enc]]
    registros = []
    for fila in filas[idx_datos:]:
        if plantilla.ignorar_filas_vacias and all(
            str(c).strip() == "" for c in fila
        ):
            continue
        registro = {}
        for i, nombre in enumerate(encabezados):
            if not nombre:
                continue
            registro[nombre] = fila[i] if i < len(fila) else ""
        registros.append(registro)
    return encabezados, registros


def detectar_columnas(archivo, plantilla: PlantillaMapeoCartola) -> list[str]:
    if plantilla.formato_archivo == PlantillaMapeoCartola.FormatoArchivo.XLSX:
        encabezados, _ = _leer_filas_xlsx(archivo, plantilla)
    else:
        encabezados, _ = _leer_filas_csv(archivo, plantilla)
    return [h for h in encabezados if h]


def aplicar_plantilla(fila: dict, plantilla: PlantillaMapeoCartola) -> dict:
    """Normaliza una fila cruda según los campos de la plantilla."""
    mapa = {c.campo_destino: c for c in plantilla.campos.all()}
    Destino = CampoMapeoCartola.CampoDestino
    out: dict[str, Any] = {}

    def valor(destino: str) -> str:
        campo = mapa.get(destino)
        if not campo:
            return ""
        raw = fila.get(campo.columna_origen, "")
        if raw is None or str(raw).strip() == "":
            return campo.valor_defecto or ""
        return str(raw).strip()

    fecha_raw = valor(Destino.FECHA_OPERACION)
    if fecha_raw:
        try:
            out["fecha_operacion"] = datetime.strptime(
                fecha_raw, plantilla.formato_fecha
            ).date()
        except ValueError as exc:
            raise ValidationError(
                f"Fecha inválida «{fecha_raw}» con formato {plantilla.formato_fecha}"
            ) from exc
    else:
        out["fecha_operacion"] = None

    out["descripcion"] = valor(Destino.DESCRIPCION)
    out["referencia"] = valor(Destino.REFERENCIA)
    out["identificador_externo"] = valor(Destino.IDENTIFICADOR_EXTERNO)
    out["contraparte"] = valor(Destino.CONTRAPARTE)

    destinos = set(mapa)
    if Destino.MONTO_INGRESO in destinos and Destino.MONTO_EGRESO in destinos:
        ingreso = _parse_monto(valor(Destino.MONTO_INGRESO), plantilla) or Decimal("0")
        egreso = _parse_monto(valor(Destino.MONTO_EGRESO), plantilla) or Decimal("0")
        if ingreso and not egreso:
            out["tipo"] = "INGRESO"
            out["monto"] = ingreso
        elif egreso and not ingreso:
            out["tipo"] = "EGRESO"
            out["monto"] = egreso
        elif ingreso and egreso:
            raise ValidationError("Cargo y abono simultáneos en la misma fila.")
        else:
            out["tipo"] = None
            out["monto"] = None
    elif Destino.TIPO in destinos and Destino.MONTO in destinos:
        out["tipo"] = valor(Destino.TIPO).upper() or None
        out["monto"] = _parse_monto(valor(Destino.MONTO), plantilla)
    else:
        monto = _parse_monto(valor(Destino.MONTO), plantilla)
        if monto is None:
            out["tipo"] = None
            out["monto"] = None
        elif monto < 0:
            out["tipo"] = "EGRESO"
            out["monto"] = abs(monto)
        else:
            out["tipo"] = "INGRESO"
            out["monto"] = monto

    saldo_raw = valor(Destino.SALDO)
    out["saldo"] = _parse_monto(saldo_raw, plantilla) if saldo_raw else None
    return out


def previsualizar_plantilla(
    *, archivo, plantilla: PlantillaMapeoCartola, limite: int = 20
) -> dict:
    """Lee un archivo de muestra y aplica el mapeo (sin crear movimientos)."""
    plantilla = (
        PlantillaMapeoCartola.objects.prefetch_related(
            Prefetch("campos", queryset=CampoMapeoCartola.objects.order_by("orden", "id"))
        ).get(pk=plantilla.pk)
    )
    validar_plantilla(plantilla=plantilla)

    if plantilla.formato_archivo == PlantillaMapeoCartola.FormatoArchivo.XLSX:
        encabezados, filas = _leer_filas_xlsx(archivo, plantilla)
    else:
        encabezados, filas = _leer_filas_csv(archivo, plantilla)

    preview = []
    errores = []
    for i, fila in enumerate(filas[: max(5, min(limite, 20))], start=1):
        try:
            preview.append(aplicar_plantilla(fila, plantilla))
        except ValidationError as exc:
            errores.append({"fila": i, "mensaje": "; ".join(exc.messages)})

    return {
        "encabezados": encabezados,
        "filas": preview,
        "errores": errores,
        "total_leidas": len(filas),
    }


# B003/B005: ver services/importacion_cartolas.py y duplicados.py
# B008: ver services/clasificacion_movimientos.py
# TODO(B009): def aplicar_movimiento(...): ...
# TODO(B011): def conciliar_parcial(...): ...
# TODO(B012): def ejecutar_reglas_cruce(...): ...
# TODO(B015): def excluir_movimiento(...): ...
# TODO(B016): def reversar_aplicacion(...): ...
# TODO(B017): def cerrar_periodo_conciliacion(...): ...
