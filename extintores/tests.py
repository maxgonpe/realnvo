from io import BytesIO
from pathlib import Path

from django.contrib.auth.models import Group, Permission, User
from django.test import SimpleTestCase
from django.test import TestCase
from django.urls import resolve, reverse
from openpyxl import load_workbook

from .models import (
    Cliente, CategoriaProducto, Intervencion, Odt, Producto, ItemIntervencion,
    EstadisticaDetalleExtintor, ImagenServicio,
)
from .views import generar_estadisticas_mensuales
from .services.stock import StockInsuficiente, ajustar_stock, guardar_consumo_item, eliminar_consumo_item
from .models import TechnicianProfile
from .forms import ItemOdtFormSet
from .permissions import (
    PERM_GESTIONAR_USUARIOS,
    PERM_FIRMAR_DOCUMENTOS,
    ROLE_TECNICO,
    puede_firmar_documentos,
    usuario_tiene_permiso,
)


class ExtintoresUrlIntegrityTests(SimpleTestCase):
    """Contratos de rutas que no requieren una base de datos."""

    def test_statistics_fixed_route_is_not_captured_as_a_month(self):
        match = resolve('/estadisticas/ver/')

        self.assertEqual(match.url_name, 'ver_estadisticas_redirect')
        self.assertEqual(match.kwargs, {})

    def test_statistics_month_route_requires_a_month(self):
        self.assertEqual(
            reverse('ver_estadisticas', kwargs={'mes': '2026-08'}),
            '/estadisticas/2026-08/',
        )

    def test_ajax_routes_have_stable_names(self):
        self.assertEqual(reverse('buscar_clientes_ajax'), '/ajax/buscar-clientes/')
        self.assertEqual(reverse('buscar_productos_ajax'), '/ajax/buscar-productos/')


class ExtintoresTemplateIntegrityTests(SimpleTestCase):
    def test_intervencion_detail_logic_is_externalized(self):
        template = (Path(__file__).parent / 'templates' / 'intervenciones' / 'detalle_intervencion.html').read_text(
            encoding='utf-8'
        )

        self.assertIn(
            "src=\"{% static 'extintores/js/detalle-intervencion.js' %}?v=1\"",
            template,
        )
        self.assertNotIn('<script>\n    // Filtro de tarjetas', template)

    def test_intervencion_formset_logic_is_externalized(self):
        template = (Path(__file__).parent / 'templates' / 'intervenciones' / 'crear.html').read_text(
            encoding='utf-8'
        )

        self.assertIn("extintores/js/intervencion-formset.js", template)
        self.assertNotIn("document.addEventListener('DOMContentLoaded'", template)

    def test_statistics_template_has_no_empty_download_links(self):
        template_path = (
            Path(__file__).parent / 'templates' / 'estadisticas' / 'generar.html'
        )
        template = template_path.read_text(encoding='utf-8')

        self.assertNotIn('href=""', template)
        self.assertNotIn('Descargar Excel', template)
        self.assertNotIn('Descargar PDF', template)

    def test_ajax_templates_do_not_hardcode_root_urls(self):
        templates = (
            Path(__file__).parent / 'templates' / 'intervenciones' / 'crear.html',
            Path(__file__).parent / 'templates' / 'intervenciones' / 'editar_consumos.html',
        )

        for template_path in templates:
            template = template_path.read_text(encoding='utf-8')
            self.assertNotIn('fetch(`/ajax/', template)

    def test_consumos_uses_external_common_and_specific_formset_modules(self):
        template = (Path(__file__).parent / 'templates' / 'intervenciones' / 'editar_consumos.html').read_text(
            encoding='utf-8'
        )
        self.assertIn("data-producto-formset", template)
        self.assertIn("data-consumo-formset", template)
        self.assertIn("extintores/js/producto-formset.js", template)
        self.assertIn("extintores/js/consumo-formset.js", template)

        common = (Path(__file__).parent / 'static' / 'extintores' / 'js' / 'producto-formset.js').read_text(
            encoding='utf-8'
        )
        specific = (Path(__file__).parent / 'static' / 'extintores' / 'js' / 'consumo-formset.js').read_text(
            encoding='utf-8'
        )
        self.assertNotIn('modal-stock-bajo', common)
        self.assertIn('modal-stock-bajo', specific)
        self.assertIn('producto-clear', common)
        self.assertIn('producto-clear', template)


class IntervencionOdtTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(nombre='Cliente de prueba')

    def test_odt_can_exist_without_intervencion(self):
        odt = Odt.objects.create()

        self.assertIsNone(odt.intervencion)

    def test_intervencion_without_odt_does_not_create_one(self):
        intervencion = Intervencion.objects.create(
            cliente=self.cliente,
            tipo='revision',
            con_odt=False,
        )

        self.assertFalse(Odt.objects.filter(intervencion=intervencion).exists())

    def test_intervencion_can_have_one_explicit_odt(self):
        intervencion = Intervencion.objects.create(
            cliente=self.cliente,
            tipo='revision',
            con_odt=True,
        )
        odt = Odt.objects.create(intervencion=intervencion)

        self.assertEqual(intervencion.odt_rel, odt)
        self.assertEqual(Odt.objects.filter(intervencion=intervencion).count(), 1)

    def test_odt_edit_formset_loads_existing_items_from_instance(self):
        odt = Odt.objects.create()
        producto = Producto.objects.create(nombre='Producto ODT', stock=5, precio_unitario=10)
        from .models import ItemOdt
        ItemOdt.objects.create(odt=odt, producto=producto, cantidad=2)

        formset = ItemOdtFormSet(instance=odt, prefix='itemodt_set')

        self.assertEqual(len(formset.forms), 1)
        self.assertEqual(formset.forms[0].instance.producto, producto)

    def test_odt_edit_template_iterates_existing_items_context(self):
        template = (Path(__file__).parent / 'templates' / 'odt' / 'editar.html').read_text(encoding='utf-8')
        self.assertIn('itemset_con_subtotales', template)
        self.assertIn('imagenes_intervencion', template)


class StockServiceTests(TestCase):
    def setUp(self):
        categoria = CategoriaProducto.objects.create(nombre='Prueba')
        self.producto = Producto.objects.create(
            nombre='Producto de prueba', categoria=categoria, stock=10
        )
        self.cliente = Cliente.objects.create(nombre='Cliente de stock')
        self.intervencion = Intervencion.objects.create(
            cliente=self.cliente, tipo='revision', alias='INT-STOCK'
        )

    def test_consumption_rejects_insufficient_stock_without_changes(self):
        with self.assertRaises(StockInsuficiente):
            ajustar_stock(self.producto.pk, -11)

        self.producto.refresh_from_db()
        self.assertEqual(self.producto.stock, 10)

    def test_consumption_and_deletion_restore_stock(self):
        item = ItemIntervencion(
            intervencion=self.intervencion, producto=self.producto, cantidad=3
        )
        guardar_consumo_item(item)
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.stock, 7)

        eliminar_consumo_item(item)
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.stock, 10)

    def test_item_model_save_does_not_change_stock_directly(self):
        item = ItemIntervencion.objects.create(
            intervencion=self.intervencion, producto=self.producto, cantidad=3
        )

        self.producto.refresh_from_db()
        self.assertEqual(self.producto.stock, 10)
        item.delete()

    def test_none_stock_is_unlimited_for_recarga_and_mantencion(self):
        for nombre in ('Recarga', 'Recarga 40%', 'Mantención'):
            categoria = CategoriaProducto.objects.create(nombre=nombre)
            producto = Producto.objects.create(nombre=nombre, categoria=categoria, stock=None)
            ajustar_stock(producto.pk, -100)
            producto.refresh_from_db()
            self.assertIsNone(producto.stock)

    def test_none_stock_is_zero_for_other_categories(self):
        categoria = CategoriaProducto.objects.create(nombre='Repuesto')
        producto = Producto.objects.create(nombre='Repuesto', categoria=categoria, stock=None)
        with self.assertRaises(StockInsuficiente):
            ajustar_stock(producto.pk, -1)

    def test_item_odt_without_price_uses_zero_subtotal(self):
        from .models import ItemOdt
        odt = Odt.objects.create()
        producto = Producto.objects.create(nombre='Servicio sin precio', stock=None)
        item = ItemOdt.objects.create(odt=odt, producto=producto, cantidad=2)
        self.assertEqual(item.precio_unitario, 0)
        self.assertEqual(item.subtotal, 0)


class PermissionTests(TestCase):
    def test_permission_can_be_granted_through_a_group(self):
        user = User.objects.create_user(username='supervisor')
        group = Group.objects.create(name='Supervisor')
        permission = Permission.objects.get(
            content_type__app_label='extintores', codename='manage_users'
        )
        group.permissions.add(permission)
        user.groups.add(group)

        self.assertTrue(usuario_tiene_permiso(user, PERM_GESTIONAR_USUARIOS))

    def test_technician_profile_grants_document_signature_without_admin_role(self):
        user = User.objects.create_user(username='tecnico')
        TechnicianProfile.objects.create(user=user)

        self.assertTrue(puede_firmar_documentos(user))

    def test_user_manager_can_assign_role_permission_and_technician_profile(self):
        manager = User.objects.create_superuser(
            username='manager', password='password123', email='manager@example.com'
        )
        target = User.objects.create_user(username='andres')
        self.client.force_login(manager)

        response = self.client.post(reverse('usuarios_simple'), {
            'action': 'editar',
            'user_id': target.pk,
            'first_name': '',
            'last_name': '',
            'email': '',
            'is_active': 'on',
            'roles': [ROLE_TECNICO],
            'permissions': [PERM_FIRMAR_DOCUMENTOS],
            'is_technician': 'on',
        })

        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertTrue(target.groups.filter(name=ROLE_TECNICO).exists())
        self.assertTrue(target.has_perm(PERM_FIRMAR_DOCUMENTOS))
        self.assertTrue(hasattr(target, 'technician_profile'))

    def test_extintores_routes_require_authentication(self):
        response = self.client.get(reverse('lista_productos'))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_authenticated_user_without_area_permission_is_rejected(self):
        user = User.objects.create_user(username='sin-permisos')
        self.client.force_login(user)

        response = self.client.get(reverse('agregar_producto'))

        self.assertEqual(response.status_code, 403)
        self.assertContains(response, 'No tienes autorización para esta sección', status_code=403)
        self.assertContains(response, 'extintores.manage_catalog', status_code=403)

    def test_user_management_button_uses_permission_not_username(self):
        template = (Path(__file__).parent / 'templates' / 'intervenciones' / 'lista.html').read_text(
            encoding='utf-8'
        )

        self.assertIn('perms.extintores.manage_users', template)
        self.assertNotIn("user.username == 'andres'", template)


class StatisticsExportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username='stats-admin', password='password123'
        )
        self.client.force_login(self.user)
        cliente = Cliente.objects.create(nombre='Cliente estadisticas')
        EstadisticaDetalleExtintor.objects.create(
            mes='2026-08', tipo_intervencion='revision', agente='PQS_40%',
            peso='6 Kg', estado='operativo', cantidad=4, cliente=cliente,
        )

    def test_excel_export_contains_month_and_quantity(self):
        response = self.client.get(reverse(
            'exportar_estadisticas_excel', kwargs={'mes': '2026-08'}
        ))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.values)
        self.assertIn('2026-08', rows[1])
        self.assertIn(4, rows[1])

    def test_pdf_export_returns_pdf(self):
        response = self.client.get(reverse(
            'exportar_estadisticas_pdf', kwargs={'mes': '2026-08'}
        ))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_grouped_pdf_returns_pdf(self):
        response = self.client.get(
            reverse('exportar_estadisticas_pdf', kwargs={'mes': '2026-08'}),
            {'agrupar': 'agente'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_invalid_month_is_rejected(self):
        response = self.client.get(reverse(
            'exportar_estadisticas_excel', kwargs={'mes': 'agosto'}
        ))

        self.assertEqual(response.status_code, 400)

    def test_excel_export_applies_dimension_filters(self):
        response = self.client.get(
            reverse('exportar_estadisticas_excel', kwargs={'mes': '2026-08'}),
            {'agente': 'PQS_40%', 'peso': '6 Kg'},
        )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.values)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][2:5], ('operativo', 'PQS_40%', '6 Kg'))

    def test_statistics_view_compares_filtered_months(self):
        EstadisticaDetalleExtintor.objects.create(
            mes='2026-07', tipo_intervencion='revision', agente='PQS_40%',
            peso='6 Kg', estado='operativo', cantidad=2,
        )
        response = self.client.get(
            reverse('ver_estadisticas', kwargs={'mes': '2026-08'}),
            {'comparar': '2026-07', 'agente': 'PQS_40%', 'peso': '6 Kg'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_actual'], 4)
        self.assertEqual(response.context['comparacion']['total'], 2)
        self.assertEqual(response.context['comparacion']['variacion'], 2)

    def test_statistics_view_groups_filtered_results(self):
        EstadisticaDetalleExtintor.objects.create(
            mes='2026-08', tipo_intervencion='revision', agente='CO2',
            peso='6 Kg', estado='operativo', cantidad=3,
        )
        response = self.client.get(
            reverse('ver_estadisticas', kwargs={'mes': '2026-08'}),
            {'agrupar': 'agente'},
        )

        self.assertEqual(response.status_code, 200)
        grouped = {row['etiqueta']: row['total'] for row in response.context['agrupacion']}
        self.assertEqual(grouped['PQS_40%'], 4)
        self.assertEqual(grouped['CO2'], 3)

    def test_grouped_excel_matches_screen_aggregation(self):
        response = self.client.get(
            reverse('exportar_estadisticas_excel', kwargs={'mes': '2026-08'}),
            {'agrupar': 'agente'},
        )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.values)
        self.assertEqual(rows[0], ('Agrupacion', 'Cantidad'))
        self.assertEqual(rows[1], ('PQS_40%', 4))

    def test_monthly_generation_is_idempotent(self):
        cliente = Cliente.objects.create(nombre='Cliente generacion')
        intervencion = Intervencion.objects.create(
            cliente=cliente, tipo='revision', fecha='2026-08-10', alias='INT-GEN'
        )
        from .models import DetalleIntervencion
        DetalleIntervencion.objects.create(
            intervencion=intervencion, agente='PQS_40%', peso='6 Kg',
            estado='operativo', presion='120'
        )

        generar_estadisticas_mensuales('2026-08')
        generar_estadisticas_mensuales('2026-08')

        detalle = EstadisticaDetalleExtintor.objects.get(mes='2026-08')
        self.assertEqual(detalle.cantidad, 1)
        self.assertEqual(EstadisticaDetalleExtintor.objects.filter(mes='2026-08').count(), 1)


class ImagenServicioTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(username='imagenes-admin', password='password123')
        self.client.force_login(self.user)
        cliente = Cliente.objects.create(nombre='Cliente imagen')
        self.intervencion = Intervencion.objects.create(
            cliente=cliente, tipo='revision', alias='INT-IMG'
        )

    def test_new_image_model_supports_order_and_description(self):
        imagen = ImagenServicio.objects.create(
            intervencion=self.intervencion,
            archivo='intervenciones/2026-08/intervencion_1/foto.jpg',
            orden=2,
            descripcion='Placa del extintor',
        )

        self.assertEqual(self.intervencion.imagenes_nuevas.get(), imagen)
        self.assertEqual(imagen.descripcion, 'Placa del extintor')

    def test_image_description_and_order_can_be_edited(self):
        imagen = ImagenServicio.objects.create(
            intervencion=self.intervencion, archivo='foto.jpg', orden=1
        )
        response = self.client.post(reverse('editar_imagen_servicio', args=[imagen.pk]), {
            'descripcion': 'Sello de seguridad', 'orden': '3',
        })

        self.assertEqual(response.status_code, 302)
        imagen.refresh_from_db()
        self.assertEqual(imagen.descripcion, 'Sello de seguridad')
        self.assertEqual(imagen.orden, 3)

    def test_image_can_be_deleted_individually(self):
        imagen = ImagenServicio.objects.create(
            intervencion=self.intervencion, archivo='foto.jpg'
        )
        response = self.client.post(reverse('eliminar_imagen_servicio', args=[imagen.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ImagenServicio.objects.filter(pk=imagen.pk).exists())


class FrontendStructureTests(TestCase):
    def test_client_autocomplete_is_external_and_keeps_named_url(self):
        template = (Path(__file__).parent / 'templates' / 'intervenciones' / 'crear.html').read_text(
            encoding='utf-8'
        )
        script = (Path(__file__).parent / 'static' / 'extintores' / 'js' / 'cliente-autocomplete.js').read_text(
            encoding='utf-8'
        )

        self.assertIn("{% static 'extintores/js/cliente-autocomplete.js' %}", template)
        self.assertIn("data-search-url=\"{% url 'buscar_clientes_ajax' %}\"", template)
        self.assertNotIn("fetch(`{% url 'buscar_clientes_ajax' %}", template)
        self.assertIn('dataset.searchUrl', script)
        self.assertIn('cliente-clear', script)

    def test_active_base_loads_external_theme_script(self):
        template = (Path(__file__).parent / 'templates' / 'base.html').read_text(
            encoding='utf-8'
        )

        self.assertIn("{% load static %}", template)
        self.assertIn("{% static 'extintores/js/theme.js' %}?v=3", template)
        self.assertIn(' defer', template)
        self.assertNotIn('document.querySelectorAll(".theme-btn")', template)
        self.assertIn('{% for message in messages %}', template)
        self.assertIn('role="alert"', template)

    def test_theme_script_limits_theme_values(self):
        script = (Path(__file__).parent / 'static' / 'extintores' / 'js' / 'theme.js').read_text(
            encoding='utf-8'
        )

        self.assertIn("['red', 'yellow', 'blue', 'gray']", script)
        self.assertIn("localStorage.setItem('theme', theme)", script)
        self.assertIn("document.readyState === 'loading'", script)
        self.assertIn("document.addEventListener('DOMContentLoaded', initialize)", script)
        self.assertIn('Use the default theme when storage is unavailable.', script)

    def test_odt_list_loads_external_search_script(self):
        template = (Path(__file__).parent / 'templates' / 'odt' / 'lista.html').read_text(encoding='utf-8')
        self.assertIn("{% static 'extintores/js/odt-search.js' %}?v=1", template)
        self.assertNotIn('new XMLHttpRequest()', template)

    def test_intervention_list_loads_external_search_script(self):
        template = (Path(__file__).parent / 'templates' / 'intervenciones' / 'lista.html').read_text(encoding='utf-8')
        self.assertIn("{% static 'extintores/js/intervencion-search.js' %}?v=1", template)
        self.assertIn("data-ajax-url=\"{% url 'ajax_intervenciones' %}\"", template)

    def test_odt_edit_loads_external_formset_script(self):
        template = (Path(__file__).parent / 'templates' / 'odt' / 'editar.html').read_text(encoding='utf-8')
        self.assertIn("{% static 'extintores/js/odt-formset.js' %}?v=1", template)
        self.assertNotIn("getElementById('add-formset').addEventListener", template)

    def test_odt_general_loads_external_formset_script(self):
        template = (Path(__file__).parent / 'templates' / 'odt' / 'editar-general.html').read_text(encoding='utf-8')
        self.assertIn("{% static 'extintores/js/odt-general-formsets.js' %}?v=1", template)
        self.assertIn('data-add-formset', template)

    def test_technician_can_operate_but_not_manage_catalog(self):
        user = User.objects.create_user(username='tecnico-ruta')
        role = Group.objects.create(name='Tecnico')
        role.permissions.set(Permission.objects.filter(
            content_type__app_label='extintores',
            codename__in=['view_operations', 'manage_operations'],
        ))
        user.groups.add(role)
        self.client.force_login(user)

        self.assertNotEqual(
            self.client.get(reverse('intervencion_lista')).status_code, 403
        )
        self.assertEqual(self.client.get(reverse('agregar_producto')).status_code, 403)

    def test_inventory_role_can_manage_stock_but_not_interventions(self):
        user = User.objects.create_user(username='inventario-ruta')
        role = Group.objects.create(name='Inventario')
        role.permissions.set(Permission.objects.filter(
            content_type__app_label='extintores',
            codename__in=['view_operations', 'manage_inventory', 'view_reports'],
        ))
        user.groups.add(role)
        self.client.force_login(user)

        self.assertNotEqual(
            self.client.get(reverse('ingreso_stock_nuevo')).status_code, 403
        )
        self.assertEqual(self.client.get(reverse('intervencion_crear')).status_code, 403)
