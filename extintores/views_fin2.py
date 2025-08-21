# === IMPORTS GENERALES ===
from datetime import timedelta
from decimal import Decimal
from collections import Counter
from collections import defaultdict
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q, F, Sum, Count
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.forms import inlineformset_factory, modelformset_factory
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string, get_template
from django.templatetags.static import static
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.timezone import now
from collections import Counter
from django.db.models import Sum
from datetime import datetime
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import ListView, CreateView, UpdateView
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
from django.shortcuts import redirect
from django.forms import modelform_factory, inlineformset_factory
from .models import IngresoStock, DetalleIngreso
from django.apps import apps
# === MODELOS ===
from .models import (
    Intervencion, DetalleIntervencion, ImagenIntervencion,
    Odt, DetalleOdt, Producto, Cliente, CompatibilidadProducto,
    ItemOdt, CategoriaProducto, Bitacora, FactorAjusteCliente,
    IngresoStock, DetalleIngreso, EstadisticaMensual,
    EstadisticaDetalleExtintor, EstadisticaDetalleProducto
)

# === FORMULARIOS ===
from .forms import (
    IntervencionForm, DetalleIntervencionForm, ItemOdtForm, ImagenIntervencionForm,
    OdtForm, DetalleOdtFormSet, ItemOdtFormSet, ProductoForm, ClienteForm,
    CategoriaForm, ImagenIntervencionFormSet, FactorAjusteClienteForm
)

# === UTILIDADES ===
from .utils import obtener_factor, descontar_stock,\
                   revertir_stock, actualizar_stock_item_odt,\
                   generar_estadisticas_mensuales

from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
import openpyxl
from django.http import HttpResponse
from .models import Producto

# === CONSTANTES ===
ESTADO_CHOICES = [
    ('operativo', 'Operativo'),
    ('f/servicio', 'Fuera/Servicio'),
    ('ph/venci.', 'PH/Vencida'),
    ('f/de/norma', 'Fuera/de/Norma'),
    ('baja/oxido', 'Baja/Oxido'),
    ('extin./abo.', 'Extintor/Abollado'),
    ('habili.+1', 'habilitado +1 año'),
    ('nuevo', 'Nuevo'),
    ('pendiente', 'Pendiente'),
    ('otro', 'Otro'),
]

# === FUNCIONES AUXILIARES ===
def registrar_bitacora(usuario, accion, modelo, objeto_id=None, descripcion=""):
    """Registra una acción en la bitácora."""
    Bitacora.objects.create(
        usuario=usuario,
        accion=accion,
        modelo=modelo,
        objeto_id=objeto_id,
        descripcion=descripcion
    )

# === VISTAS ODT ===

class OdtListView(ListView):
    model = Odt
    template_name = 'odt/lista.html'
    context_object_name = 'odts'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('intervencion__cliente', 'tecnico').order_by('-id')
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(
                Q(alias__icontains=search) |
                Q(tecnico__first_name__icontains=search) |
                Q(tecnico__last_name__icontains=search) |
                Q(intervencion__cliente__nombre__icontains=search)
            )
        return queryset

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('odt/includes/odt_resultados.html', context)
            return JsonResponse({'html': html})
        return super().render_to_response(context, **response_kwargs)

def eliminar_odt(request, pk):
    odt = get_object_or_404(Odt, pk=pk)
    if request.method == 'POST':
        odt.delete()
        registrar_bitacora(
            usuario=request.user,
            accion='Eliminar',
            modelo='Odt',
            objeto_id=pk,
            descripcion=f"El usuario {request.user.username} eliminó la Odt #{odt.pk} con fecha {odt.fecha.strftime('%Y-%m-%d')}."
        )
        return redirect('odt_lista')
    return render(request, 'odt/eliminar.html', {'odt': odt})


def editar_odt(request, pk):
    odt = get_object_or_404(Odt, pk=pk)
    productos_disponibles = Producto.objects.select_related('categoria').order_by('categoria__nombre', 'nombre')

    if request.method == 'POST':
        form = OdtForm(request.POST, instance=odt)
        formset = DetalleOdtFormSet(request.POST, instance=odt, prefix='detalleodt')
        #itemset = ItemOdtFormSet(request.POST, queryset=odt.items.all(), prefix='itemodt_set')
        itemset = ItemOdtFormSet(request.POST or None, instance=odt, prefix='itemodt_set')

        if form.is_valid() and formset.is_valid() and itemset.is_valid():
            form.save()
            formset.save()

            original_items = {item.pk: item for item in odt.items.all()}

            items = itemset.save(commit=False)
            for item in items:
                item.odt = odt
                item.save()
                item_original = original_items.get(item.pk)
                actualizar_stock_item_odt(item, item_original)

            for obj in itemset.deleted_objects:
                revertir_stock(obj.producto, obj.cantidad)
                obj.delete()

            return redirect('odt_lista')
        else:
            print("Errores form:", form.errors)
            print("Errores detalle formset:", formset.errors)
            print("Errores itemset:", itemset.errors)

    else:
        form = OdtForm(instance=odt)
        formset = DetalleOdtFormSet(instance=odt, prefix='detalleodt')
        itemset = ItemOdtFormSet(queryset=odt.items.all(), prefix='itemodt_set')

    componentes = {
        'exterior': 'exterior',
        'valvula': 'Valvula',
        'manometro': 'Manometro',
        'manguera': 'Manguera',
        'cintillo': 'Cintillo',
        'agente_extintor': 'Agente Extintor',
        'tubo_sifon': 'Tubo Sifon',
        'sellos_seguro': 'Sellos de Seguridad',
        'seguridad': 'Sistema de Seguridad',
        'rotulacion': 'Rotulacion',
        'recarga': 'recarga',
        'mantencion': 'mantencion',
        'correcta': 'Instalación Correcta',
        'acceso': 'Acceso',
        'instrucciones': 'Instrucciones'
    }

    sugerencias = []
    for detalle in odt.detalles.all():
        problemas = []
        for campo, nombre in componentes.items():
            if getattr(detalle, campo) is False:
                problemas.append(nombre)
        if problemas:
            sugerencias.append({'extintor': detalle, 'problemas': problemas})

    registrar_bitacora(
        usuario=request.user,
        accion='Editar',
        modelo='Odt',
        objeto_id=pk,
        descripcion=f"El usuario {request.user.username} editó la Odt #{odt.pk} con fecha {odt.fecha.strftime('%Y-%m-%d')}."
    )

    return render(request, 'odt/editar.html', {
        'form': form,
        'formset': formset,
        'itemset': itemset,
        'odt': odt,
        'sugerencias': sugerencias,
        'productos_disponibles': productos_disponibles
    })



def odt_listar(request):
    odts = Odt.objects.all().order_by('-id')
    return render(request, 'odt/lista.html', {'odts': odts})

def odt_agregar_productos(request, pk):
    odt = get_object_or_404(Odt, pk=pk)
    productos_disponibles = Producto.objects.all().order_by('categoria__nombre', 'nombre')

    componentes = {
        'exterior': 'Exterior',
        'valvula': 'Valvula',
        'manometro': 'Manometro',
        'manguera': 'Manguera',
        'cintillo': 'Cintillo',
        'agente_extintor': 'Agente Extintor',
        'tubo_sifon': 'Tubo Sifon',
        'sellos_seguro': 'Sellos de Seguridad',
        'seguridad': 'Sistema de Seguridad',
        'rotulacion': 'Rotulacion',
        'recarga': 'Recarga',
        'mantencion': 'Mantencion',
        'correcta': 'Instalación Correcta',
        'acceso': 'Acceso',
        'instrucciones': 'Instrucciones'
    }

    # --- Generar sugerencias por extintor ---
    sugerencias = []
    for detalle in odt.detalles.all():
        problemas = []
        for campo, nombre in componentes.items():
            if getattr(detalle, campo) is False:
                problemas.append({'campo': campo, 'nombre': nombre})
        if problemas:
            sugerencias.append({
                'extintor': detalle,
                'problemas': problemas,
                'agente': detalle.agente,
                'peso': detalle.peso,
            })

    # Ordenar sugerencias: por nombre del primer problema, agente y peso
    sugerencias.sort(
        key=lambda x: (
            x['problemas'][0]['nombre'] if x['problemas'] else '',
            x['agente'] or '',
            x['peso'] or 0,
        )
    )

    # --- Agrupar por problema individual ---
    temp_agrupado = defaultdict(list)
    for sug in sugerencias:
        for problema in sug['problemas']:
            temp_agrupado[problema['nombre']].append({
                'extintor': sug['extintor'],
                'agente': sug['agente'],
                'peso': sug['peso'],
                'campo': problema['campo'],
                'nombre': problema['nombre'],
            })

    sugerencias_agrupadas = dict(temp_agrupado)
    # --- Procesar formulario POST ---
    if request.method == 'POST':
        for sug in sugerencias:
            detalle = sug['extintor']
            for problema in sug['problemas']:
                campo = problema['campo']
                key_sel = f'seleccionado_{detalle.pk}_{campo}'
                key_prod = f'producto_id_{detalle.pk}_{campo}'
                key_cant = f'cantidad_{detalle.pk}_{campo}'
                if request.POST.get(key_sel) and request.POST.get(key_prod):
                    producto_id = request.POST.get(key_prod)
                    cantidad = int(request.POST.get(key_cant, 1))
                    producto = Producto.objects.get(pk=producto_id)

                    item, created = ItemOdt.objects.get_or_create(
                        odt=odt,
                        producto=producto,
                        defaults={'cantidad': cantidad}
                    )
                    if not created:
                        diferencia = cantidad - item.cantidad
                        item.cantidad = cantidad
                        item.save()
                        if diferencia > 0:
                            descontar_stock(producto, diferencia)
                    else:
                        descontar_stock(producto, cantidad)

        return redirect('odt_agregar_productos', pk=odt.pk)

    # --- Calcular total ---
    total = sum(item.subtotal for item in odt.items.all())

    # --- Resumen agrupado para mostrar arriba ---
    resumen = Counter()
    for sug in sugerencias:
        agente = sug['agente'] or 'N/A'
        peso = sug['peso'] or 'N/A'
        for problema in sug['problemas']:
            clave = (agente, peso, problema['nombre'])
            resumen[clave] += 1

    resumen_agrupado = [
        {'agente': k[0], 'peso': k[1], 'problema': k[2], 'cantidad': v}
        for k, v in resumen.items()
    ]
    resumen_agrupado.sort(key=lambda x: (x['problema'], x['agente'], x['peso']))

    # --- Render final ---
    print("DEBUG - sugerencias count:", len(sugerencias))
    print("DEBUG - sugerencias_agrupadas keys:", list(sugerencias_agrupadas.keys()))

    return render(request, 'odt/agregar_productos.html', {
        'odt': odt,
        'productos_disponibles': productos_disponibles,
        'sugerencias': sugerencias,
        'sugerencias_agrupadas': sugerencias_agrupadas,
        'total': total,
        'resumen_agrupado': resumen_agrupado,
    })



def odt_editar_items(request, pk):
    odt = get_object_or_404(Odt, pk=pk)
    queryset = ItemOdt.objects.filter(odt=odt)

    ItemOdtFormSet = modelformset_factory(
        ItemOdt,
        fields=('producto', 'cantidad'),
        extra=0,
        can_delete=True
    )

    formset = ItemOdtFormSet(request.POST or None, queryset=queryset)

    if request.method == 'POST':
        if formset.is_valid():
            original_items = {item.pk: item for item in queryset}

            instances = formset.save(commit=False)
            for instance in instances:
                instance.odt = odt
                instance.save()

                item_original = original_items.get(instance.pk)
                actualizar_stock_item_odt(instance, item_original)

            for obj in formset.deleted_objects:
                revertir_stock(obj.producto, obj.cantidad)
                obj.delete()

            return redirect('odt_editar_items', pk=odt.pk)
        else:
            print("Errores:", formset.errors)

    total = sum(item.subtotal for item in queryset)

    return render(request, 'odt/editar_items.html', {
        'odt': odt,
        'formset': formset,
        'total': total,
    })


def odt_detalle(request, pk):
    odt = get_object_or_404(Odt, pk=pk)
    productos_disponibles = Producto.objects.all().order_by('categoria__nombre', 'nombre')
    
    componentes = {
        'exterior': 'Exterior',
        'valvula': 'Valvula',
        'manometro': 'Manometro',
        'manguera': 'Manguera',
        'cintillo': 'Cintillo',
        'agente_extintor': 'Agente Extintor',
        'tubo_sifon': 'Tubo Sifon',
        'sellos_seguro': 'Sellos de Seguridad',
        'seguridad': 'Sistema de Seguridad',
        'rotulacion': 'Rotulacion',
        'recarga': 'Recarga',
        'mantencion': 'Mantencion',
        'revision': 'Revision',
        'correcta': 'Instalación Correcta',
        'acceso': 'Acceso',
        'instrucciones': 'Instrucciones'
    }

    # Generar sugerencias con claves únicas basadas en campo
    sugerencias = []
    for detalle in odt.detalles.all():
        problemas = []
        for campo, nombre in componentes.items():
            if getattr(detalle, campo) is False:
                problemas.append({'campo': campo, 'nombre': nombre})
        if problemas:
            sugerencias.append({
                'extintor': detalle,
                'problemas': problemas,
            })

    # Configuración del Formset
    ItemOdtFormSet = modelformset_factory(
        ItemOdt,
        fields=('producto', 'cantidad'),
        extra=0,
        can_delete=True
    )
    
    # Procesamiento de formularios
    if request.method == 'POST':
        # Formulario de sugerencias
        if 'agregar_sugeridos' in request.POST:
            for sug in sugerencias:
                detalle = sug['extintor']
                for problema in sug['problemas']:
                    campo = problema['campo']
                    key_sel = f'seleccionado_{detalle.pk}_{campo}'
                    key_prod = f'producto_id_{detalle.pk}_{campo}'
                    key_cant = f'cantidad_{detalle.pk}_{campo}'
                    
                    if request.POST.get(key_sel) and request.POST.get(key_prod):
                        producto_id = request.POST.get(key_prod)
                        cantidad = int(request.POST.get(key_cant, 1))
                        producto = Producto.objects.get(pk=producto_id)
                        
                        item, created = ItemOdt.objects.get_or_create(
                            odt=odt,
                            producto=producto,
                            defaults={'cantidad': cantidad}
                        )
                        if not created:
                            item.cantidad = cantidad
                            item.save()
            return redirect('odt_detalle', pk=odt.pk)
        
        # Formulario de edición/eliminación
        elif 'guardar_cambios' in request.POST:
            formset = ItemOdtFormSet(
                request.POST, 
                queryset=ItemOdt.objects.filter(odt=odt)
            )
            
            if formset.is_valid():
                instances = formset.save(commit=False)
                for instance in instances:
                    instance.odt = odt
                    instance.save()
                
                for obj in formset.deleted_objects:
                    obj.delete()
                
                return redirect('odt_detalle', pk=odt.pk)
    
    # Inicialización GET o POST no válido
    formset = ItemOdtFormSet(queryset=ItemOdt.objects.filter(odt=odt))
    
    total = sum(item.subtotal for item in odt.items.all())

    return render(request, 'odt/detalle.html', {
        'odt': odt,
        'productos_disponibles': productos_disponibles,
        'sugerencias': sugerencias,
        'formset': formset,
        'total': total,
    })


def odt_pdf(request, pk):
    odt = get_object_or_404(Odt, pk=pk)
    items = odt.items.all()
    total = sum(item.subtotal for item in items)
    iva = total * Decimal('0.19')
    total_final = total + iva
    html_content = render_to_string("odt/pdf_odt.html", {
        "odt": odt,
        "items": items,
        "total": total,
        "iva": iva,
        "total_final": total_final,
    })
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = f'inline; filename="ODT-{odt.pk}.pdf"'
    HTML(string=html_content,base_url=request.build_absolute_uri('/')).write_pdf(response)
    # ... registrar_bitacora ...
    registrar_bitacora(
        usuario=request.user,
        accion='Crear',
        modelo='Odt',
        objeto_id=pk,
        descripcion=f"El usuario {request.user.username} creo el documento PDF para la Odt #{odt.pk} con fecha {odt.fecha.strftime('%Y-%m-%d')}."
    )


    return response


def odt_excel(request, pk):
    odt = Odt.objects.get(pk=pk)
    wb = Workbook()
    ws = wb.active
    ws.title = f"ODT #{odt.pk}"

    ws.append(["ODT #", odt.pk])
    ws.append(["Fecha", str(odt.fecha)])
    ws.append(["Cliente", odt.intervencion.cliente.nombre])
    ws.append(["Tipo", odt.intervencion.get_tipo_display()])
    ws.append([])

    ws.append(["Detalles de Extintores"])
    ws.append(["Ubicación", "Agente", "Presión", "PH", "Estado"])
    for d in odt.intervencion.detalles.all():
        ws.append([d.ubicacion, d.agente, d.presion, d.ph_ultima, d.estado])
    ws.append([])

    ws.append(["Productos/Servicios"])
    ws.append(["Producto", "Categoría", "Cantidad", "Precio Unitario", "Subtotal"])
    total = 0
    for item in odt.items.all():
        subtotal = item.subtotal
        total += subtotal
        ws.append([
            item.producto.nombre,
            #item.producto.categoria,
            item.producto.categoria.nombre if item.producto.categoria else "",
            item.cantidad,
            float(item.producto.precio_unitario),
            float(subtotal)
        ])
    ws.append(["", "", "", "Total", float(total)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"ODT-{odt.pk}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    registrar_bitacora(
                        usuario=request.user,
                        accion='Crear',
                        modelo='Odt',
                        objeto_id=pk,
                        descripcion = f"El usuario {request.user.username} creo el documento Excel para la Odt #{odt.pk} con fecha {odt.fecha.strftime('%Y-%m-%d')}."
                        
                    )

    return response

# === VISTAS INTERVENCIÓN ===

class IntervencionListView(LoginRequiredMixin, ListView):
    model = Intervencion
    template_name = 'intervenciones/lista.html'
    context_object_name = 'intervenciones'
    def get_queryset(self):
        queryset = Intervencion.objects.all().order_by('-id')
        query = self.request.GET.get("q", "").strip()
        if query:
            queryset = queryset.filter(
                Q(cliente__nombre__icontains=query) |
                Q(alias__icontains=query) |
                Q(tecnico__first_name__icontains=query) |
                Q(tecnico__last_name__icontains=query) |
                Q(tipo__icontains=query)
            )
        return queryset


class IntervencionAjaxListView(LoginRequiredMixin, ListView):
    model = Intervencion
    context_object_name = 'intervenciones'

    def get_queryset(self):
        query = self.request.GET.get("q", "").strip()
        queryset = Intervencion.objects.all().order_by('-id')
        if query:
            queryset = queryset.filter(
                Q(cliente__nombre__icontains=query) |
                Q(alias__icontains=query) |
                Q(tecnico__first_name__icontains=query) |
                Q(tecnico__last_name__icontains=query) |
                Q(tipo__icontains=query)
                # No se puede buscar por fecha con icontains si es un campo Date
            )
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        html = render_to_string(
            "intervenciones/partials/intervenciones_list.html",
            {"intervenciones": queryset}
        )
        return JsonResponse({"html": html})

def crear_intervencion(request):
    prefix = 'detalles'
    
    # Definir el formset aquí para usarlo consistentemente
    DetalleIntervencionFormSet = inlineformset_factory(
        Intervencion,
        DetalleIntervencion,
        form=DetalleIntervencionForm,
        extra=0,  # Cero formularios iniciales
        can_delete=True,
        validate_min=False
    )
    
    if request.method == 'POST':
        form = IntervencionForm(request.POST)
        formset = DetalleIntervencionFormSet(request.POST, prefix=prefix)
        imagenes_form = ImagenIntervencionForm(request.POST, request.FILES)
        
        # Validar TODOS los formularios juntos
        if all([form.is_valid(), formset.is_valid(), imagenes_form.is_valid()]):
            with transaction.atomic():
                # Guardar la intervención principal
                intervencion = form.save(commit=False)
                intervencion.save()  # Necesario para tener PK antes de relaciones
                
                # Guardar el formset de detalles
                instances = formset.save(commit=False)
                for instance in instances:
                    # Ignorar formularios vacíos (sin agente)
                    if not instance.agente:
                        continue
                    instance.intervencion = intervencion
                    instance.save()
                
                # Manejar objetos marcados para borrar
                for obj in formset.deleted_objects:
                    obj.delete()
                
                # Guardar las imágenes
                imagen_instance = imagenes_form.save(commit=False)
                imagen_instance.intervencion = intervencion
                imagen_instance.save()
                
                # Crear ODT si aplica
                if intervencion.con_odt:
                    odt = Odt.objects.create(
                        intervencion=intervencion,
                        tecnico=intervencion.tecnico,
                        alias=intervencion.alias,
                        observaciones=intervencion.notas
                    )
                    
                    for d in intervencion.detalles.all():
                        DetalleOdt.objects.create(
                            odt=odt,
                            agente=d.agente,
                            nro_precinto=d.nro_precinto,
                            ubicacion=d.ubicacion,
                            sello_incen=d.sello_incen,
                            sello_incen_ste=d.sello_incen_ste,
                            certificado=d.certificado,
                            estado=d.estado,
                            exterior=d.exterior,
                            ph_ultima=d.ph_ultima,
                            ph_estado=d.ph_estado,
                            ph_vencimiento=d.ph_vencimiento,
                            valvula=d.valvula,
                            manometro=d.manometro,
                            manguera=d.manguera,
                            cintillo=d.cintillo,
                            agente_extintor=d.agente_extintor,
                            tubo_sifon=d.tubo_sifon,
                            sellos_seguro=d.sellos_seguro,
                            seguridad=d.seguridad,
                            rotulacion=d.rotulacion,
                            ultima_fecha=d.ultima_fecha,
                            recarga=d.recarga,
                            mantencion=d.mantencion,
                            presion=d.presion,
                            peso=d.peso,
                            correcta=d.correcta,
                            acceso=d.acceso,
                            instrucciones=d.instrucciones,
                            #habilitado_un_ano=d.habilitado_un_ano,
                            #baja_por_oxido=d.baja_por_oxido,
                            #baja_por_ph=d.baja_por_ph,
                            #baja_por_fuera_norma=d.baja_por_fuera_norma,
                        )

                    registrar_bitacora(
                        usuario=request.user,
                        accion='Crear',
                        modelo='Odt',
                        objeto_id=odt.pk,
                        descripcion=f"El usuario {request.user.username} creó la ODT #{odt.pk} con fecha {odt.fecha.strftime('%Y-%m-%d')}."
                    )
                
                return redirect('intervencion_lista')
        else:
            # Manejo de errores detallado
            print("Formulario no válido")
            print("Errores en IntervencionForm:", form.errors)
            print("Errores en Formset:", formset.errors)
            print("Errores en ImagenIntervencionForm:", imagenes_form.errors)
            
            # Renderizar nuevamente con errores
            return render(request, 'intervenciones/crear.html', {
                'form': form,
                'formset': formset,
                'imagenes_form': imagenes_form,
                'prefix': prefix
            })
    
    else:  # GET request
        form = IntervencionForm()
        # Iniciar con 0 formularios (queryset vacío)
        formset = DetalleIntervencionFormSet(
            prefix=prefix, 
            queryset=DetalleIntervencion.objects.none()
        )
        imagenes_form = ImagenIntervencionForm()

        return render(request, 'intervenciones/crear.html', {
            'form': form,
            'formset': formset,
            'imagenes_form': imagenes_form,
            'prefix': prefix
        })



def editar_intervencion(request, pk):
    intervencion = get_object_or_404(Intervencion, pk=pk)
    imagenes_instance = ImagenIntervencion.objects.filter(intervencion=intervencion).first()
    
    DetalleIntervencionFormSet = inlineformset_factory(
        Intervencion,
        DetalleIntervencion,
        form=DetalleIntervencionForm,
        extra=1,
        can_delete=True,
        validate_min=False
    )

    if request.method == 'POST':
        form = IntervencionForm(request.POST, instance=intervencion)
        formset = DetalleIntervencionFormSet(request.POST, instance=intervencion, prefix='detalles')
        imagenes_form = ImagenIntervencionForm(request.POST, request.FILES, instance=imagenes_instance)

        if form.is_valid() and formset.is_valid() and imagenes_form.is_valid():
            intervencion = form.save()

            # Guardar el formset (pero sin commit para modificar)
            instances = formset.save(commit=False)

            for instance in instances:
                if not instance.pk and not instance.agente:
                    continue

                # 👉 Forzamos estado si hay una causa de baja
                #if instance.baja_por_oxido or instance.baja_por_ph or instance.baja_por_fuera_norma:
                #    instance.estado = "f/servicio"
                #if instance.habilitado_un_ano:
                #    instance.estado = "habili.+1"
                #if instance.baja_por_oxido:
                #    instance.estado = "baja/oxido"
                #elif instance.baja_por_ph:
                #    instance.estado = "baja/ph"
                #elif instance.baja_por_fuera_norma:
                #    instance.estado = "f/de/norma"
                #elif instance.extintor_abollado:
                #    instance.estado = "extin./abo."
                #elif instance.baja_otro:
                #    instance.estado = "f/servicio"
                
                                    

                instance.intervencion = intervencion
                instance.save()

            # Guardamos eliminaciones
            for obj in formset.deleted_objects:
                obj.delete()

            # ✅ Guardar todos los detalles restantes para que estén en la DB
            formset.save_m2m()

            # Guardar imagen
            imagen_instance = imagenes_form.save(commit=False)
            imagen_instance.intervencion = intervencion
            imagen_instance.save()

            # ✅ Ahora sí, detalles guardados — podemos crear la ODT
            if intervencion.con_odt and not hasattr(intervencion, 'odt_rel'):
                odt = Odt.objects.create(intervencion=intervencion, tecnico=intervencion.tecnico, alias=intervencion.alias, observaciones=intervencion.notas)
                for d in intervencion.detalles.all():
                    DetalleOdt.objects.create(
                        odt=odt,
                        agente=d.agente,
                        nro_precinto=d.nro_precinto,
                        ubicacion=d.ubicacion,
                        sello_incen=d.sello_incen,
                        sello_incen_ste=d.sello_incen_ste,
                        certificado=d.certificado,
                        estado=d.estado,
                        exterior=d.exterior,
                        ph_ultima=d.ph_ultima,
                        ph_estado=d.ph_estado,
                        ph_vencimiento=d.ph_vencimiento,
                        valvula=d.valvula,
                        manometro=d.manometro,
                        manguera=d.manguera,
                        cintillo=d.cintillo,
                        agente_extintor=d.agente_extintor,
                        tubo_sifon=d.tubo_sifon,
                        sellos_seguro=d.sellos_seguro,
                        seguridad=d.seguridad,
                        rotulacion=d.rotulacion,
                        ultima_fecha=d.ultima_fecha,
                        recarga=d.recarga,
                        mantencion=d.mantencion,
                        presion=d.presion,
                        peso=d.peso,
                        correcta=d.correcta,
                        acceso=d.acceso,
                        instrucciones=d.instrucciones,
                        #habilitado_un_ano=d.habilitado_un_ano,
                        #baja_por_oxido=d.baja_por_oxido,
                        #baja_por_ph=d.baja_por_ph,
                        #baja_por_fuera_norma=d.baja_por_fuera_norma,
                    )

                registrar_bitacora(
                    usuario=request.user,
                    accion='Crear',
                    modelo='Odt',
                    objeto_id=intervencion.pk,
                    descripcion=f"El usuario {request.user.username} generó una ODT para la intervención #{intervencion.pk}."
                )

            return redirect('intervencion_detalle', pk=intervencion.pk)
        else:
            print("Formulario no válido")
            print("Errores en IntervencionForm:", form.errors)
            print("Errores en Formset:", formset.errors)
            print("Errores en ImagenIntervencionForm:", imagenes_form.errors)
    else:
        form = IntervencionForm(instance=intervencion)
        formset = DetalleIntervencionFormSet(instance=intervencion, prefix='detalles')
        imagenes_form = ImagenIntervencionForm(instance=imagenes_instance)

    context = {
        'form': form,
        'formset': formset,
        'imagenes_form': imagenes_form,
        'intervencion': intervencion,
    }
    return render(request, 'intervenciones/editar_intervencion.html', context)



def eliminar_intervencion(request, pk):
    intervencion = get_object_or_404(Intervencion, pk=pk)
    if request.method == 'POST':
        intervencion.delete()
        return redirect('intervencion_lista')  # Cambia por la URL de tu lista de productos

    registrar_bitacora(
                usuario=request.user,
                accion='Eliminar una Intervención',
                modelo='Intervencion',
                objeto_id=pk,
                descripcion = f"El usuario {request.user.username} eliminó la intervención #{intervencion.pk} con fecha {intervencion.fecha.strftime('%Y-%m-%d')}."
                
            )

    return render(request, 'intervenciones/eliminar.html', {'intervencion': intervencion})


def detalle_intervencion2(request, pk):
    intervencion = get_object_or_404(Intervencion, pk=pk)
    detalles = intervencion.detalles.all()
    # Contar por estado
    conteo = Counter([d.estado for d in detalles])
    # Contar por agente
    conteo_agente = Counter([d.agente for d in detalles])
    print('conteo -->',conteo)
    print(conteo_agente)  # Para ver qué contiene
    print(type(conteo_agente))  # Para verificar que es un Counter

    # Contar por peso
    conteo_peso = Counter([d.peso for d in detalles])
    print(conteo_peso)  # Para ver qué contiene
    print(type(conteo_peso))  # Para verificar que es un Counter

    estadisticas_boleanas = {
        'Habilitado por un año': detalles.filter(habilitado_un_ano=True).count(),
        'Baja por óxido': detalles.filter(baja_por_oxido=True).count(),
        'Baja por pH': detalles.filter(baja_por_ph=True).count(),
        'Baja por fuera de norma': detalles.filter(baja_por_fuera_norma=True).count(),
        'Extintor abollado': detalles.filter(extintor_abollado=True).count(),
        'Baja por otro motivo': detalles.filter(baja_otro=True).count(),
        'Extintor nuevo': detalles.filter(extintor_nuevo=True).count(),
    }


    #context['estadisticas_boleanas'] = estadisticas_boleanas
    context = {
    'intervencion': intervencion,
    'detalles': detalles,
    'conteo': conteo,
    'conteo_agente': conteo_agente,
    'conteo_peso': conteo_peso,
    'estadisticas_boleanas': estadisticas_boleanas,
}


    estadistica = []
    total = 0
    for key, label in ESTADO_CHOICES:
        cantidad = conteo.get(key, 0)
        estadistica.append({'estado': label, 'cantidad': cantidad})
        total += cantidad

    # Resumen agrupado de problemas (¡AHORA SÍ antes del return!)
    componentes = {
        'exterior': 'Exterior',
        'valvula': 'Válvula',
        'manometro': 'Manómetro',
        'manguera': 'Manguera',
        'cintillo': 'Cintillo',
        'agente_extintor': 'Agente Extintor',
        'tubo_sifon': 'Tubo Sifon',
        'sellos_seguro': 'Sellos de Seguridad',
        'seguridad': 'Sistema de Seguridad',
        'rotulacion': 'Rotulación',
        'recarga': 'Recarga',
        'mantencion': 'Mantención',
        'correcta': 'Instalación Correcta',
        'acceso': 'Acceso',
        'instrucciones': 'Instrucciones'
    }

    resumen = Counter()
    for d in detalles:
        for campo, nombre in componentes.items():
            if getattr(d, campo) is False:
                clave = (d.agente or 'N/A', d.peso or 'N/A', nombre)
                resumen[clave] += 1

    resumen_agrupado = [
        {'agente': k[0], 'peso': k[1], 'problema': k[2], 'cantidad': v}
        for k, v in resumen.items()
    ]
    resumen_agrupado.sort(key=lambda x: (x['problema'], x['agente'], x['peso']))

    estadisticas_estado = {
    dict(ESTADO_CHOICES).get(k, k): v for k, v in conteo.items()
    }

    return render(request, 'intervenciones/detalle_intervencion.html', {
        'intervencion': intervencion,
        'estadistica': estadistica,
        'total_extintores': total,
        'resumen_agrupado': resumen_agrupado,  # <- ya visible en el template
        'conteo_agente': dict(conteo_agente),  # Pasar conteo por agente
        'conteo_peso': dict(conteo_peso),      # Pasar conteo por peso
        'estadisticas_estado': dict(estadisticas_estado),
        'estadisticas_boleanas': dict(estadisticas_boleanas),
    })

from collections import Counter

def detalle_intervencion(request, pk):
    intervencion = get_object_or_404(Intervencion, pk=pk)
    detalles = intervencion.detalles.all()

    # Contar por estado
    conteo = Counter([d.estado for d in detalles])

    # Contar por agente
    conteo_agente = Counter([d.agente for d in detalles])

    # Contar por peso
    conteo_peso = Counter([d.peso for d in detalles])

    # Contar por combinación agente + peso
    conteo_agente_peso = Counter([
        (d.agente or "N/A", d.peso or "N/A")
        for d in detalles
    ])

    estadisticas_boleanas = {
        'Habilitado por un año': detalles.filter(habilitado_un_ano=True).count(),
        'Baja por óxido': detalles.filter(baja_por_oxido=True).count(),
        'Baja por pH': detalles.filter(baja_por_ph=True).count(),
        'Baja por fuera de norma': detalles.filter(baja_por_fuera_norma=True).count(),
        'Extintor abollado': detalles.filter(extintor_abollado=True).count(),
        'Baja por otro motivo': detalles.filter(baja_otro=True).count(),
        'Extintor nuevo': detalles.filter(extintor_nuevo=True).count(),
    }

    # Estadística por estado
    estadistica = []
    total = 0
    for key, label in ESTADO_CHOICES:
        cantidad = conteo.get(key, 0)
        estadistica.append({'estado': label, 'cantidad': cantidad})
        total += cantidad

    # Resumen agrupado de problemas
    componentes = {
        'exterior': 'Exterior',
        'valvula': 'Válvula',
        'manometro': 'Manómetro',
        'manguera': 'Manguera',
        'cintillo': 'Cintillo',
        'agente_extintor': 'Agente Extintor',
        'tubo_sifon': 'Tubo Sifon',
        'sellos_seguro': 'Sellos de Seguridad',
        'seguridad': 'Sistema de Seguridad',
        'rotulacion': 'Rotulación',
        'recarga': 'Recarga',
        'mantencion': 'Mantención',
        'correcta': 'Instalación Correcta',
        'acceso': 'Acceso',
        'instrucciones': 'Instrucciones'
    }

    resumen = Counter()
    for d in detalles:
        for campo, nombre in componentes.items():
            if getattr(d, campo) is False:
                clave = (d.agente or 'N/A', d.peso or 'N/A', nombre)
                resumen[clave] += 1

    resumen_agrupado = [
        {'agente': k[0], 'peso': k[1], 'problema': k[2], 'cantidad': v}
        for k, v in resumen.items()
    ]
    resumen_agrupado.sort(key=lambda x: (x['problema'], x['agente'], x['peso']))

    estadisticas_estado = {
        dict(ESTADO_CHOICES).get(k, k): v for k, v in conteo.items()
    }

    return render(request, 'intervenciones/detalle_intervencion.html', {
        'intervencion': intervencion,
        'estadistica': estadistica,
        'total_extintores': total,
        'resumen_agrupado': resumen_agrupado,
        'conteo_agente': dict(conteo_agente),
        'conteo_peso': dict(conteo_peso),
        'conteo_agente_peso': dict(conteo_agente_peso),  # 🔹 nuevo conteo combinado
        'estadisticas_estado': dict(estadisticas_estado),
        'estadisticas_boleanas': dict(estadisticas_boleanas),
    })



class IntervencionExcel(View):
    def get(self, request, pk):
        intervencion = Intervencion.objects.get(pk=pk)
        detalles = intervencion.detalles.all().order_by('id')

        wb = Workbook()
        ws = wb.active

        # Encabezado
        ws['A1'] = f'PLANILLA DE {intervencion.get_tipo_display().upper()}'
        ws['A2'] = f'Cliente: {intervencion.cliente}'
        ws['C2'] = f'Técnico: {intervencion.tecnico}'
        ws['E2'] = f'Fecha: {intervencion.fecha}'
        ws.merge_cells('A1:H1')

        # Títulos
        headers = [
            'N°', 'Agente','Nro Precinto','Ubicación', 'Sello INCEN','Sello INCEN STE', 'Certificado','Estado', 'Cilindro  Exterior','Ultima PH', 'Estado PH',
            'Vencimiento PH','Valvula','Manometro','Manguera','Cintillo','Agente Extintor','Tubo Sifon', 'Sellos Seguro', 'Dispositivo Seguridad', 'Rotulacion',
            'Ult Fecha Mantto.', 'Recarga', 'Mantencion', 'Presion', 'Peso', 'Inst Correcta', 'Acceso', 'Instrucciones'
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=4, column=col_num).value = header

        # Estilos para vencido
        #fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        #fuente_roja = Font(color="9C0006")
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # habilitado_un_ano
        fill_gris = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")   # baja_por_oxido
        fill_amarillo = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # baja_por_ph
        fuente_blanca = Font(color="FFFFFF")
        fuente_roja = Font(color="9C0006")  # usada antes para vencido

        # Filas
        row = 5
        for idx, d in enumerate(detalles, 1):
            fila = [
                idx,
                d.agente,
                d.nro_precinto,
                d.ubicacion,
                d.sello_incen,
                d.sello_incen_ste,
                d.certificado,
                d.estado,
                "OK" if d.exterior else "X",
                d.ph_ultima,
                d.ph_estado,
                d.ph_vencimiento,
                "OK" if d.valvula else "X",
                "OK" if d.manguera else "X",
                "OK" if d.manometro else "X",
                "OK" if d.cintillo else "X",
                "OK" if d.agente_extintor else "X",
                "OK" if d.tubo_sifon else "X",
                "OK" if d.sellos_seguro else "X",
                "OK" if d.seguridad else "X",
                "OK" if d.rotulacion else "X",
                d.ultima_fecha.strftime('%Y-%m-%d') if d.ultima_fecha else '',
                "OK" if d.recarga else "X",
                "OK" if d.mantencion else "X",
                d.presion,
                d.peso,
                "OK" if d.correcta else "X",
                "OK" if d.acceso else "X",
                "OK" if d.instrucciones else "X",
            ]

            for col_num, valor in enumerate(fila, 1):
                celda = ws.cell(row=row, column=col_num, value=valor)

                # Si contiene "vencido" en ph_estado (columna 11), aplicar formato
                #if d.ph_estado and 'vencido' in d.ph_estado.lower():
                #    celda.fill = fill_verde
                #    celda.font = fuente_roja

                # Aplicar formato por ph vencido (columna 11)
                if d.ph_estado and 'vencido' in d.ph_estado.lower():
                    celda.fill = fill_verde
                    celda.font = fuente_roja

                # Aplicar formato por tipo de estado general
                if d.baja_por_oxido:
                    celda.fill = fill_gris
                    celda.font = fuente_blanca
                elif d.baja_por_ph:
                    celda.fill = fill_amarillo
                elif d.baja_por_fuera_norma:
                    celda.fill = fill_amarillo
                elif d.habilitado_un_ano:
                    celda.fill = fill_verde

            row += 1

        nombre_archivo = f"{intervencion.get_tipo_display()}-{intervencion.pk}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        wb.save(response)

        registrar_bitacora(
                        usuario=self.request.user,
                        accion='Crear',
                        modelo='Intervencion',
                        objeto_id=pk,
                        descripcion = f"El usuario {request.user.username} creo el documento Excel para el Servicio #{intervencion.pk} con fecha {intervencion.fecha.strftime('%Y-%m-%d')}."
                        
                    )

        return response


class IntervencionPDF(View):
    def get(self, request, pk):
        intervencion = Intervencion.objects.select_related(
            'tecnico', 'tecnico__technician_profile'
        ).get(pk=pk)

        detalles = intervencion.detalles.all().order_by('id')
        imagenes = intervencion.imagenes.first()

        # Estadística por estado
        conteo_estado = Counter([d.estado for d in detalles])
        estadistica = []
        total = 0
        for key, label in ESTADO_CHOICES:
            cantidad = conteo_estado.get(key, 0)
            estadistica.append({'estado': label, 'cantidad': cantidad})
            total += cantidad

        # Conteo por agente
        conteo_agente = Counter([d.agente for d in detalles if d.agente])
        
        # Conteo por peso
        conteo_peso = Counter([d.peso for d in detalles if d.peso])

        html_string = render_to_string('intervenciones/pdf_intervencion.html', {
            'intervencion': intervencion,
            'detalles': detalles,
            'imagenes': imagenes,
            'estadistica': estadistica,
            'total_extintores': total,
            'conteo_agente': dict(conteo_agente),
            'conteo_peso': dict(conteo_peso),
        })

        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = html.write_pdf(stylesheets=[CSS(string='@page { size: A4; margin: 1cm }')])

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Intervencion-{pk}.pdf"'

        registrar_bitacora(
            usuario=request.user,
            accion='Crear',
            modelo='Intervencion',
            objeto_id=pk,
            descripcion=f"El usuario {request.user.username} creó el PDF para el Servicio #{intervencion.pk} del {intervencion.fecha.strftime('%Y-%m-%d')}."
        )

        return response

# === VISTAS PRODUCTO ===


def lista_productos(request):
    productos = Producto.objects.all().order_by('nombre').order_by('categoria')
    return render(request, 'producto/lista.html', {'productos': productos})


def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()

            registrar_bitacora(
                        usuario=request.user,
                        accion='Agregar',
                        modelo='Productos',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Agrego el producto: ({form.cleaned_data['nombre']}) de la categoria, ({form.cleaned_data['categoria']}) con precio de ({form.cleaned_data['precio_unitario']})  en el catalogo, con fecha {timezone.now().strftime('%Y-%m-%d')}."
                                                
                    )


            return redirect('lista_productos')
    else:
        form = ProductoForm()
    return render(request, 'producto/form.html', {'form': form, 'titulo': 'Nuevo Producto'})


def modificar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()

            registrar_bitacora(
                        usuario=request.user,
                        accion='Modificar',
                        modelo='Productos',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Modifico el producto: ({form.cleaned_data['nombre']}) de la categoria, ({form.cleaned_data['categoria']}) con precio de ({form.cleaned_data['precio_unitario']})  en el catalogo, con fecha {timezone.now().strftime('%Y-%m-%d')}."
                        
                    )

            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'producto/form.html', {'form': form, 'titulo': 'Editar Producto'})


def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':

        registrar_bitacora(
            usuario=request.user,
            accion='Eliminar',
            modelo='Producto',
            objeto_id=pk,
            descripcion=(
                f"El usuario {request.user.username} eliminó el producto '{producto.nombre}' "
                f"(ID: {pk}), categoría: '{producto.categoria}', precio: ${producto.precio_unitario}, "
                f"en fecha {timezone.now().strftime('%Y-%m-%d')}."
            )
        )

        producto.delete()
        
        return redirect('lista_productos')
    return render(request, 'producto/eliminar.html', {'producto': producto})

# === VISTAS CLIENTE ===

def lista_clientes(request):
    clientes = Cliente.objects.all().order_by('nombre')
    return render(request, 'cliente/lista.html', {'clientes': clientes})


def agregar_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():

            registrar_bitacora(
                        usuario=request.user,
                        accion='Agregar',
                        modelo='Cliente',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Agrego un Nuevo cliente: ({form.cleaned_data['nombre']}) con rut, ({form.cleaned_data['rut']}) y contacto ({form.cleaned_data['contacto']}), con fecha {timezone.now().strftime('%Y-%m-%d')}."
                                                
                    )

            form.save()
            return redirect('lista_clientes')
    else:
        form = ClienteForm()
    return render(request, 'cliente/form.html', {'form': form, 'titulo': 'Nuevo Cliente'})

def modificar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()

            registrar_bitacora(
                        usuario=request.user,
                        accion='Modificar',
                        modelo='Cliente',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Modifico el registro del Cliente: ({form.cleaned_data['nombre']}) con el rut, ({form.cleaned_data['rut']}) con el contacto ({form.cleaned_data['contacto']}), con fecha {timezone.now().strftime('%Y-%m-%d')}."
                        
                    )


            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'cliente/form.html', {'form': form, 'titulo': 'Editar Cliente'})

def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':

        registrar_bitacora(
                        usuario=request.user,
                        accion='Eliminar',
                        modelo='Cliente',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Elimino el cliente: ({cliente.nombre}) con fecha {timezone.now().strftime('%Y-%m-%d')}."
                                                
                    )

        cliente.delete()
        return redirect('lista_clientes')
    return render(request, 'cliente/eliminar.html', {'cliente': cliente})



# === VISTAS CATEGORÍA ===

def lista_categorias(request):
    categorias = CategoriaProducto.objects.all().order_by('nombre')
    return render(request, 'categoria/lista.html', {'categorias': categorias})

def agregar_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()

            registrar_bitacora(
                        usuario=request.user,
                        accion='Crear',
                        modelo='Categoria',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Creo la categoria: ({form.cleaned_data['nombre']}) con fecha {timezone.now().strftime('%Y-%m-%d')}."
                                                
                    )

            return redirect('lista_categorias')
    else:
        form = CategoriaForm()
    return render(request, 'categoria/form.html', {'form': form, 'titulo': 'Nueva Categoria'})

def modificar_categoria(request, pk):
    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()

            registrar_bitacora(
                        usuario=request.user,
                        accion='Crear',
                        modelo='Categoria',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Modifico la categoria: ({form.cleaned_data['nombre']}) con fecha {timezone.now().strftime('%Y-%m-%d')}."
                                                
                    )

            return redirect('lista_categorias')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'categoria/form.html', {'form': form, 'titulo': 'Editar Categoria'})

def eliminar_categoria(request, pk):
    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    if request.method == 'POST':

        registrar_bitacora(
                        usuario=request.user,
                        accion='Eliminar',
                        modelo='Categoria',
                        objeto_id=None,
                        descripcion = f"El usuario {request.user.username} Elimino la categoria: ({pk}) con fecha {timezone.now().strftime('%Y-%m-%d')}."
                                                
                    )


        categoria.delete()
        return redirect('lista_categorias')
    return render(request, 'categoria/eliminar.html', {'categoria': categoria})


# === VISTAS FACTOR AJUSTE CLIENTE ===


def factorajustecliente_lista(request):
    factores = FactorAjusteCliente.objects.select_related('cliente', 'categoria').order_by('cliente__nombre', 'categoria__nombre')
    return render(request, 'factor/factorajustecliente_lista.html', {'factores': factores})

def factorajustecliente_crear(request):
    if request.method == 'POST':
        form = FactorAjusteClienteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('factorajustecliente_lista')
    else:
        form = FactorAjusteClienteForm()
    return render(request, 'factor/factorajustecliente_form.html', {'form': form, 'titulo': 'Nuevo Factor de Ajuste'})

def factorajustecliente_editar(request, pk):
    factor = get_object_or_404(FactorAjusteCliente, pk=pk)
    if request.method == 'POST':
        form = FactorAjusteClienteForm(request.POST, instance=factor)
        if form.is_valid():
            form.save()
            return redirect('factorajustecliente_lista')
    else:
        form = FactorAjusteClienteForm(instance=factor)
    return render(request, 'factor/factorajustecliente_form.html', {'form': form, 'titulo': 'Editar Factor de Ajuste'})

def factorajustecliente_eliminar(request, pk):
    factor = get_object_or_404(FactorAjusteCliente, pk=pk)
    if request.method == 'POST':
        factor.delete()
        return redirect('factorajustecliente_lista')
    return render(request, 'factor/factorajustecliente_eliminar.html', {'factor': factor})


@require_POST
def agregar_item_odt(request, odt_pk):
    detalle = get_object_or_404(DetalleOdt, pk=request.POST.get('detalle_id'))
    producto = get_object_or_404(Producto, pk=request.POST.get('producto_id'))
    cantidad = int(request.POST.get('cantidad', 1))

    # Crear item asociado a la ODT
    ItemOdt.objects.create(
        odt=detalle.odt,
        producto=producto,
        cantidad=cantidad
    )

    # Crear compatibilidad genérica
    CompatibilidadProducto.objects.get_or_create(
        producto=producto,
        detalle_odt=detalle,
        defaults={'motivo': 'Sugerencia automática desde componente defectuoso'}
    )

    return redirect('odt_editar', pk=odt_pk)

IngresoStockForm = modelform_factory(IngresoStock, fields=['observaciones'])

DetalleIngresoFormSet = inlineformset_factory(
    IngresoStock,
    DetalleIngreso,
    fields=('producto', 'cantidad'),
    extra=1,
    can_delete=True
)

def ingreso_stock_nuevo(request):
    form = IngresoStockForm(request.POST or None)
    formset = DetalleIngresoFormSet(request.POST or None, prefix='detalleingreso_set')

    if request.method == 'POST':
        if form.is_valid() and formset.is_valid():
            ingreso = form.save()
            detalles = formset.save(commit=False)
            for d in detalles:
                d.ingreso = ingreso
                d.save()
            return redirect('lista_productos')  # o redirige donde prefieras

    return render(request, 'producto/ingreso_form.html', {
        'form': form,
        'formset': formset
    })


def exportar_inventario_excel(request):
    accesorios = Producto.objects.filter(categoria__nombre__iexact='Accesorio')
    #accesorios = Producto.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Accesorios"

    # Encabezados
    ws.append(['Nombre', 'Categoría', 'Precio Unitario', 'Stock', 'Valor Total'])
    total_valor = 0

    for producto in accesorios:
        valor_total = (producto.stock or 0) * (producto.precio_unitario or 0)
        total_valor += valor_total 
        ws.append([
            producto.nombre,
            producto.categoria.nombre if producto.categoria else 'Sin categoría',
            float(producto.precio_unitario or 0),
            float(producto.stock or 0),
            float(valor_total)
        ])

    # Línea de total
    ws.append([])
    ws.append(['', '', '', 'TOTAL INVENTARIO:', total_valor])

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Inventario_Accesorios.xlsx'
    wb.save(response)
    return response


def exportar_inventario_pdf(request):
    accesorios = Producto.objects.filter(categoria__nombre='Accesorio')
    #accesorios = Producto.objects.all()
    total_valor = 0
    for p in accesorios:
        p.valor_total = (p.stock or 0) * (p.precio_unitario or 0)
        total_valor += p.valor_total

    html_string = render_to_string('producto/inventario_pdf.html', {
        'productos': accesorios,
        'total_valor': total_valor
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename=Inventario_Accesorios.pdf'

    with tempfile.NamedTemporaryFile(delete=True) as output:
        HTML(string=html_string).write_pdf(output.name)
        output.seek(0)
        response.write(output.read())

    return response




def generar_estadisticas_mensuales(mes=None):
    print("Entrando en la función de los extintores")
    if mes is None:
        mes = now().strftime('%Y-%m')

    fecha_inicio = datetime.strptime(mes, "%Y-%m")
    fecha_fin = datetime(fecha_inicio.year + int(fecha_inicio.month == 12), 
                         fecha_inicio.month % 12 + 1, 1)

    # Inicializar contadores para productos
    total_productos_utilizados = 0

    # Recopilar estadísticas de intervenciones
    intervenciones_queryset = Intervencion.objects.filter(
        fecha__gte=fecha_inicio,
        fecha__lt=fecha_fin
    ).select_related('cliente')

    for intervencion in intervenciones_queryset:
        cantidad_extintores = DetalleIntervencion.objects.filter(
            intervencion=intervencion
        ).count()

        EstadisticaMensual.objects.update_or_create(
            mes=mes,
            tipo='intervencion',
            cliente=intervencion.cliente,
            tipo_intervencion=intervencion.tipo if hasattr(intervencion, 'tipo') else 'default',
            defaults={
                'cantidad_intervenciones': 1,
                'cantidad_extintores': cantidad_extintores
            }
        )

    # Recopilar estadísticas de ODTs
    odts_queryset = Odt.objects.filter(
        fecha__gte=fecha_inicio,
        fecha__lt=fecha_fin
    ).select_related('intervencion__cliente')

    for odt in odts_queryset:
        EstadisticaMensual.objects.update_or_create(
            mes=mes,
            tipo='odt',
            cliente=odt.intervencion.cliente,
            defaults={
                'cantidad_odt': 1
            }
        )


    # Recopilar estadísticas de productos utilizados en ODTs
    item_queryset = ItemOdt.objects.filter(
        odt__fecha__gte=fecha_inicio,
        odt__fecha__lt=fecha_fin
    ).select_related('producto', 'odt__intervencion__cliente')

    productos_dict = {}

    for entry in item_queryset.values('producto', 'producto__nombre', 'odt__intervencion__cliente', 'cantidad', 'producto__categoria'):
        producto_id = entry['producto']
        cliente_id = entry.get('odt__intervencion__cliente')
        cantidad = entry['cantidad']
        nombre = entry['producto__nombre']
        categoria = entry.get('producto__categoria')  # Asegúrate de tener acceso a la categoría

        if cliente_id is None:
            continue

        total_productos_utilizados += cantidad  # Acumular el total de productos utilizados

        if (producto_id, cliente_id) not in productos_dict:
            productos_dict[(producto_id, cliente_id)] = {
                'cantidad': 0,
                'nombre': nombre,
                'categoria': categoria
            }

        productos_dict[(producto_id, cliente_id)]['cantidad'] += cantidad

    for (producto_id, cliente_id), data in productos_dict.items():
        precio_unitario = ItemOdt.objects.filter(producto_id=producto_id).first().precio_unitario
        subtotal = data['cantidad'] * precio_unitario if precio_unitario else 0

        EstadisticaDetalleProducto.objects.update_or_create(
            mes=mes,
            producto_id=producto_id,
            cliente_id=cliente_id,
            defaults={
                'cantidad': data['cantidad'],
                'subtotal': subtotal,
                'nombre': data['nombre']
            }
        )

    # Llenar estadísticas de extintores tratados
    detalle_queryset = DetalleIntervencion.objects.filter(
        intervencion__fecha__gte=fecha_inicio,
        intervencion__fecha__lt=fecha_fin
    ).select_related('intervencion')

    
    # Agrupar los detalles por combinación única de campos
    conteo = defaultdict(int)

    for entry in detalle_queryset:
        tipo_intervencion = entry.intervencion.tipo if hasattr(entry.intervencion, 'tipo') else 'default'
        cliente_id = entry.intervencion.cliente.id if hasattr(entry.intervencion, 'cliente') else None

        key = (
            tipo_intervencion,
            entry.agente,
            entry.peso,
            entry.estado,
            cliente_id,
        )

        conteo[key] += 1  # contamos cuántas veces aparece esa combinación

    # Guardar en la base de datos
    for (tipo_intervencion, agente, peso, estado, cliente_id), cantidad in conteo.items():
        EstadisticaDetalleExtintor.objects.update_or_create(
            mes=mes,
            tipo_intervencion=tipo_intervencion,
            agente=agente,
            peso=peso,
            estado=estado,
            cliente_id=cliente_id,
            defaults={'cantidad': cantidad}
        )


    # Actualizar la estadística mensual con los totales
    for cliente_id in productos_dict.keys():
        estadistica = EstadisticaMensual.objects.filter(mes=mes, tipo='productos', cliente=cliente_id).first()
        if estadistica:
            estadistica.total_productos_utilizados += total_productos_utilizados
            estadistica.categoria_producto = productos_dict[cliente_id]['categoria']  # Asignar la categoría
            estadistica.save()


@login_required
def generar_estadisticas_view(request):
    if request.method == 'POST':
        mes = request.POST.get('mes')
        if mes:
            generar_estadisticas_mensuales(mes)
            messages.success(request, f'Estadísticas generadas para el mes {mes}.')
            return redirect('ver_estadisticas', mes=mes)

    # Obtener meses disponibles desde Intervenciones y ODTs
    Intervencion = apps.get_model('extintores', 'Intervencion')
    Odt = apps.get_model('extintores', 'Odt')

    fechas_intervenciones = list(Intervencion.objects.values_list('fecha', flat=True))
    fechas_odt = list(Odt.objects.values_list('fecha', flat=True))

    todos_meses = sorted(
        set(dt.strftime('%Y-%m') for dt in fechas_intervenciones + fechas_odt if dt),
        reverse=True
    )

    mes_actual = todos_meses[0] if todos_meses else now().strftime("%Y-%m")

    # Agrupación por estado
    estados_interes = [
        'operativo', 'f/servicio', 'ph/venci.', 'f/de/norma',
        'baja/oxido', 'extin./abo.', 'habili.+1', 'nuevo'
    ]

    estadisticas_por_estado = (
        EstadisticaDetalleExtintor.objects
        .filter(mes=mes_actual, estado__in=estados_interes)
        .values('estado')
        .annotate(total=Count('cantidad'))
        .order_by('estado')
    )

    estadisticas_estado_dict = {estado: 0 for estado in estados_interes}
    for item in estadisticas_por_estado:
        estado = item['estado']
        total = item['total']
        estadisticas_estado_dict[estado] = total

    # Conteo real por estado (sin filtro de estados_interes)
    conteo_extintores_estado = (
        EstadisticaDetalleExtintor.objects
        .filter(mes=mes_actual)
        .values('estado')
        .annotate(cantidad=Sum('cantidad'))
        .order_by('estado')
    )

    conteo_extintores_dict = {
        item['estado']: item['cantidad'] for item in conteo_extintores_estado
    }

    return render(request, 'estadisticas/generar.html', {
        'meses_disponibles': todos_meses,
        'mes_actual': mes_actual,
        'estadisticas_por_estado': estadisticas_estado_dict,
        'conteo_extintores_estado': conteo_extintores_dict,
        
    })



@login_required
def ver_estadisticas_view(request, mes=None):
    if mes is None:
        mes = now().strftime("%Y-%m")

    # 1) Totales generales
    cantidad_intervenciones = EstadisticaMensual.objects.filter(
        mes=mes, tipo='intervencion'
    ).aggregate(total=Sum('cantidad_intervenciones'))['total'] or 0

    cantidad_extintores = EstadisticaMensual.objects.filter(
        mes=mes, tipo='intervencion'
    ).aggregate(total=Sum('cantidad_extintores'))['total'] or 0

    cantidad_odt = EstadisticaMensual.objects.filter(
        mes=mes, tipo='odt'
    ).aggregate(total=Sum('cantidad_odt'))['total'] or 0

    cantidad_ventas = EstadisticaMensual.objects.filter(
        mes=mes, tipo_intervencion='venta'
    ).aggregate(total=Sum('cantidad_extintores'))['total'] or 0

    cantidad_revision = EstadisticaMensual.objects.filter(
        mes=mes, tipo='intervencion', tipo_intervencion='revision'
    ).aggregate(total=Sum('cantidad_extintores'))['total'] or 0

    cantidad_mantencion = EstadisticaMensual.objects.filter(
        mes=mes, tipo='intervencion', tipo_intervencion='mantencion'
    ).aggregate(total=Sum('cantidad_extintores'))['total'] or 0

    cantidad_recarga = EstadisticaMensual.objects.filter(
        mes=mes, tipo='intervencion', tipo_intervencion='recarga'
    ).aggregate(total=Sum('cantidad_extintores'))['total'] or 0

    # 2) Estadísticas por estado (bloque para el nuevo cuadro)
    estados_a_contar = [
        'operativo', 'f/servicio', 'ph/venci.', 'f/de/norma',
        'baja/oxido', 'extin./abo.', 'habili.+1', 'nuevo'
    ]

    conteo_por_estado = EstadisticaDetalleExtintor.objects.filter(
        mes=mes,
        estado__in=estados_a_contar
    ).values('estado').annotate(total=Sum('cantidad'))

    estadisticas_estado = {estado: 0 for estado in estados_a_contar}
    for fila in conteo_por_estado:
        estadisticas_estado[fila['estado']] = fila['total']

    # 3) Detalle de extintores por estado + agente + peso
    extintores = EstadisticaDetalleExtintor.objects.filter(
        mes=mes
    ).values('estado', 'agente', 'peso').annotate(
        cantidad=Sum('cantidad')
    ).order_by('estado', 'agente', 'peso')

    # 4) Totales por estado
    totales_por_estado = EstadisticaDetalleExtintor.objects.filter(
        mes=mes
    ).values('estado').annotate(
        total=Sum('cantidad')
    )
    total_dict = {item['estado']: item['total'] for item in totales_por_estado}

    # 5) Productos vendidos
    productos = EstadisticaDetalleProducto.objects.filter(mes=mes)

    return render(request, "estadisticas/ver.html", {
        "mes": mes,
        "productos": productos,
        "extintores": extintores,
        "total_dict": total_dict,
        "resumen_datos": {
            "intervenciones": cantidad_intervenciones,
            "extintores": cantidad_extintores,
            "odts": cantidad_odt,
            "ventas": cantidad_ventas,
            "revision": cantidad_revision,
            "mantencion": cantidad_mantencion,
            "recarga": cantidad_recarga,
        },
        "estadisticas_estado": estadisticas_estado,
    })

